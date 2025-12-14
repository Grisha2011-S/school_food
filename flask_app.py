from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, g, send_from_directory, send_file
import os
import time
import logging
import io
import binascii
from pathlib import Path
from logging.handlers import RotatingFileHandler
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy import func

# Минимальная инициализация приложения до определения роутов
app = Flask(__name__, static_folder='static', template_folder='templates')
# Загружаем конфиг из instance/config.py если есть
try:
    app.config.from_pyfile('instance/config.py')
except Exception:
    # если файла нет — достаточно конфигурации по умолчанию
    pass
app.secret_key = app.config.get('SECRET_KEY', os.environ.get('SECRET_KEY', 'dev-secret'))

# Декоратор проверки авторизации (определён заранее, чтобы использовать в декораторах маршрутов)
from functools import wraps
def login_required(role=None):
    """Декоратор проверки авторизации, размещён рядом с инициализацией app.

    Позволяет использовать @login_required(...) выше по файлу при определении роутов.
    """
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            # если нет user_id в сессии — перенаправляем на общий логин
            if 'user_id' not in session:
                flash('Сначала войдите в систему.', 'error')
                return redirect(url_for('login'))

            if role:
                if isinstance(role, (list, tuple, set)):
                    allowed = set(role)
                else:
                    allowed = {role}
                if session.get('role') not in allowed:
                    flash('Доступ запрещён!', 'error')
                    return redirect(url_for('login'))

            return f(*args, **kwargs)
        return decorated_function
    return decorator
from datetime import timedelta, datetime, date
from flask_caching import Cache
from models import db, Parents, Student, Cook, Eat, EatLog
from models import City, School, Grade, Admin, Pack, PackItem
from admin_utils import verify_admin, create_admin, activate_admin
from nutrition_calc import validate_measurements, calculate_nutrition
# Установим безопасные значения конфигурации по умолчанию для локального запуска
instance_dir = os.path.join(os.path.dirname(__file__), 'instance')
os.makedirs(instance_dir, exist_ok=True)

# По умолчанию используем локальную SQLite-базу в каталоге instance
app.config.setdefault('SQLALCHEMY_DATABASE_URI', 'sqlite:///' + os.path.join(instance_dir, 'school_food.db'))
app.config.setdefault('SQLALCHEMY_TRACK_MODIFICATIONS', False)

# Включим простой кэш по умолчанию (файловый/простой in-memory) если не настроен
app.config.setdefault('CACHE_TYPE', 'SimpleCache')

# Конфигурация куков для совместимости с современными браузерами
app.config['SESSION_COOKIE_SECURE'] = True  # Отправлять только по HTTPS
app.config['SESSION_COOKIE_HTTPONLY'] = True  # Недоступно для JavaScript
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'  # Защита от CSRF
app.config['PERMANENT_SESSION_LIFETIME'] = 86400 * 7  # 7 дней

# Инициализируем кеш сразу после создания app, чтобы декораторы @cache.* работали
cache = Cache()
cache.init_app(app)
# Пример: GEMINI_API_KEY = "ваш_секретный_ключ"
GEMINI_API_KEY = "AIzaSyALiST2y0go1Aen3_GnJjzjbr6UsBevn3I"
# Если вы явно укажете ключ в переменной ниже, он будет экспортирован в окружение
# и подхвачен функцией analyze_image_with_gemini из food_detection_impl.py
# if GEMINI_API_KEY:
#     try:
#         os.environ['GEMINI_API_KEY'] = GEMINI_API_KEY
#         app.logger.info('GEMINI_API_KEY configured from flask_app variable')
#     except Exception:
#         app.logger.exception('Failed to set GEMINI_API_KEY in environment')


def _ensure_analyze_image_loaded() -> bool:
    """Ленивая загрузка реализации анализа изображения.

    При первом вызове пытается импортировать `analyze_image_with_gemini`
    из `food_detection` (который в свою очередь лениво подхватывает
    реализацию из `food_detection_impl`). Результат записывается в
    `app._analyze_image`. Возвращает True, если функция доступна.
    """
    # Если уже проверяли — вернуть результат
    if getattr(app, '_analyze_image_checked', False):
        return bool(getattr(app, '_analyze_image', None))

    try:
        from food_detection import analyze_image_with_gemini
        app._analyze_image = analyze_image_with_gemini
        app._analyze_image_checked = True
        app.logger.info('Loaded analyze_image_with_gemini')
        return True
    except Exception as e:
        app._analyze_image = None
        app._analyze_image_checked = True
        app.logger.warning(f'Image analysis implementation not available: {e}')
        return False

# Языковые настройки
import json
from functools import wraps

def load_translations(lang):
    """Загружает переводы для указанного языка"""
    import logging
    logger = logging.getLogger(__name__)
    
    # Абсолютный путь к папке с переводами (используем pathlib)
    translations_dir = BASE_DIR / 'translations'
    
    try:
        # Сначала пробуем загрузить запрошенный язык
        path = translations_dir / f'{lang}.json'
        if path.exists():
            with open(path, 'r', encoding='utf-8') as f:
                translations = json.load(f)
                logger.info(f"Loaded translations for '{lang}' from {path}")
                return translations
                
        # Если не нашли, пробуем fallback на русский
        fallback = translations_dir / 'ru.json'
        if fallback.exists():
            with open(fallback, 'r', encoding='utf-8') as f:
                translations = json.load(f)
                logger.warning(f"Translation file for '{lang}' not found, using fallback ru.json")
                return translations
                
        # Если и русский не нашли
        logger.error(f"No translations found for '{lang}' and no fallback ru.json in {translations_dir}")
        return {}
        
    except Exception as e:
        logger.exception(f"Error loading translations for {lang}: {str(e)}")
        return {}

# Декоратор для добавления переводов в контекст шаблона
def with_translations(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        g.lang = session.get('language', 'ru')  # По умолчанию русский
        if not hasattr(g, 'translations') or not g.translations:
            g.translations = load_translations(g.lang)
        return f(*args, **kwargs)
    return decorated_function

# Функция проверки флешки
def check_admin_flash_drive():
    """
    Проверяет наличие специальной флешки для админа.
    Флешка должна содержать файл admin_key.txt с правильным ключом.
    """
    try:
        # Проверяем наличие флешки (Windows)
        import win32file
        app.logger.info("=== Начинаем проверку флешек ===")

        # Ожидаемый ключ админа
        ADMIN_KEY = "admin123"  # Замените на ваш секретный ключ

        # Получаем битовую маску подключенных дисков
        drives = win32file.GetLogicalDrives()
        app.logger.info(f"Обнаружена битовая маска дисков: {bin(drives)}")

        # Перебираем все возможные буквы дисков (A-Z)
        for i in range(26):
            if drives & (1 << i):
                drive_letter = chr(65 + i) + ':\\'
                try:
                    # Определяем тип диска
                    drive_type = win32file.GetDriveType(drive_letter)
                    app.logger.info(f"Диск {drive_letter} - Тип: {drive_type} "
                                  f"(REMOVABLE={win32file.DRIVE_REMOVABLE}, "
                                  f"FIXED={win32file.DRIVE_FIXED}, "
                                  f"REMOTE={win32file.DRIVE_REMOTE}, "
                                  f"CDROM={win32file.DRIVE_CDROM}, "
                                  f"RAMDISK={win32file.DRIVE_RAMDISK})")

                    # Проверяем, является ли диск съемным (флешкой)
                    if drive_type == win32file.DRIVE_REMOVABLE:
                        app.logger.info(f"Найден съемный диск: {drive_letter}")

                        # Проверяем наличие файла ключа
                        key_path = os.path.join(drive_letter, 'admin_key.txt')
                        app.logger.info(f"Проверяем наличие файла: {key_path}")

                        if os.path.exists(key_path):
                            app.logger.info(f"Файл найден: {key_path}")
                            try:
                                with open(key_path, 'r', encoding='utf-8') as f:
                                    key = f.read().strip()
                                    app.logger.info(f"Файл прочитан, длина ключа: {len(key)}")

                                    # Сравниваем с ожидаемым ключом
                                    if key == ADMIN_KEY:
                                        app.logger.info("=== Найдена флешка администратора с верным ключом ===")
                                        return ADMIN_KEY

                            except Exception as e:
                                app.logger.error(f"Ошибка при чтении файла {key_path}: {str(e)}")
                        else:
                            app.logger.info(f"Файл не найден: {key_path}")

                except Exception as e:
                    app.logger.error(f"Ошибка при проверке диска {drive_letter}: {str(e)}")
                    continue

        app.logger.info("=== Проверка завершена, флешка админа не найдена ===")
        return False

    except ImportError:
        app.logger.error("Модуль win32file не установлен. Установите его с помощью: pip install pywin32")
        return False
    except Exception as e:
        app.logger.error(f"Критическая ошибка при проверке флешки: {str(e)}")
        return False

# Проверка наличия необходимых модулей
try:
    import win32file
except ImportError:
    app.logger.error("Модуль win32file не установлен. Установите его с помощью: pip install pywin32")

# Регистрируем декоратор для всех роутов
@app.before_request
def before_request():
    # Skip heavy checks for static assets and simple GETs that don't need DB
    try:
        path = request.path or ''
    except Exception:
        path = ''
    if path.startswith('/static') or path in ('/favicon.ico', '/robots.txt'):
        return
    # Приоритет: session -> cookie -> по умолчанию 'ru'
    lang = session.get('language')
    if not lang:
        # fallback: если сессия не сохранилась (например, Secure cookie при http),
        # попробуем прочитать значение из обычного cookie и записать в сессию
        cookie_lang = request.cookies.get('language')
        if cookie_lang in ('ru', 'kk'):
            session['language'] = cookie_lang
            session.permanent = True
            lang = cookie_lang
    
    app.logger.info(f"Текущая роль: {session.get('role')}")
            
    # Проверяем наличие админа в базе при первом запуске (обёрнуто в try/except)
    if not hasattr(app, '_admin_checked'):
        try:
            app.logger.info("Проверка наличия админа в базе...")
            admin = None
            try:
                admin = Admin.query.first()
            except Exception:
                # Если БД недоступна — не ломаем обработку запроса (особенно для статических файлов)
                app.logger.warning('Database not available yet while checking admin')

            if not admin:
                app.logger.info("Создаем первого админа...")
                admin_login = "admin"
                admin_password = "admin123"  # Начальный пароль
                try:
                    admin = Admin(
                        login=admin_login,
                        password=generate_password_hash(admin_password),
                        is_master=True
                    )
                    db.session.add(admin)
                    db.session.commit()
                    app.logger.info(f"Создан первый админ с логином: {admin_login} и паролем: {admin_password}")
                    print(f"\n\nСоздан первый админ:\nЛогин: {admin_login}\nПароль: {admin_password}\n\n")
                except Exception as e:
                    app.logger.error(f"Ошибка при создании админа: {str(e)}")
                    db.session.rollback()
        except Exception as e:
            app.logger.exception('Unexpected error in admin check')
        finally:
            app._admin_checked = True
        
    current_lang = lang or 'ru'
    g.lang = current_lang
    
    # Загружаем переводы только если их нет или язык изменился
    if not hasattr(g, 'translations') or not g.translations or getattr(g, 'current_lang', None) != current_lang:
        g.translations = load_translations(current_lang)
        g.current_lang = current_lang
        app.logger.debug(f"Loaded translations for language: {current_lang}")

# Добавляем функцию перевода в контекст шаблона
@app.context_processor
def utility_processor():
    def translate(key):
        """
        Получает перевод по ключу.
        Поддерживает вложенные ключи через точку.
        Например: 
        {{ t('auth.login') }} - получит translations['auth']['login']
        {{ t('welcome') }} - получит translations['welcome']
        """
        if not hasattr(g, 'translations'):
            g.translations = {}

        translations = g.translations
        parts = key.split('.')
        current = translations
        
        for part in parts:
            if isinstance(current, dict) and part in current:
                current = current[part]
            else:
                # Если перевод не найден, пробуем загрузить файл переводов и повторить попытку
                g.translations = load_translations(g.lang if hasattr(g, 'lang') else 'ru')
                translations = g.translations
                current = translations

                # Повторная попытка получить перевод по всем частям
                found = True
                for p in parts:
                    if isinstance(current, dict) and p in current:
                        current = current[p]
                    else:
                        found = False
                        break

                if not found:
                    # логируем отсутствующий ключ и возвращаем пустую строку
                    app.logger.debug(f"Missing translation for key: {key} (lang={g.lang})")
                    return ''

        return current if current is not None else ''

    # expose role display helper to templates
    # get_role_display можно вызывать в шаблонах как {{ get_role_display('cook') }}
    return dict(t=translate, get_role_display=get_role_display)


# Утилита: читаемое имя роли для отображения в шаблонах
def get_role_display(role_key: str) -> str:
    """Возвращает человеко-читаемое имя роли по ключу.

    Поддерживает ключи: 'student', 'parent', 'cook', 'admin' и любые другие — возвращает исходный ключ.
    """
    if not role_key:
        return ''
    mapping = {
        'student': 'Ученик',
        'parent': 'Родитель',
        'cook': 'Диетсестра',
        'admin': 'Администратор',
        'teacher': 'Учитель'
    }
    return mapping.get(str(role_key), str(role_key))

# Роут для изменения языка
@app.route('/lang/<lang>')
def change_lang(lang):
    if lang in ['ru', 'kk']:
        session['language'] = lang
        session.permanent = True
        flash(f"Language set: {lang}")
    # Формируем redirect-ответ и устанавливаем простой cookie 'language' как backup
    resp = redirect(request.referrer or url_for('index'))
    try:
        secure = bool(app.config.get('SESSION_COOKIE_SECURE', False))
        samesite = app.config.get('SESSION_COOKIE_SAMESITE', None)
        max_age = int(app.permanent_session_lifetime.total_seconds())
        if samesite:
            resp.set_cookie('language', lang, max_age=max_age, secure=secure, httponly=False, samesite=samesite)
        else:
            resp.set_cookie('language', lang, max_age=max_age, secure=secure, httponly=False)
    except Exception:
        # если что-то пошло не так при установке cookie — всё равно делаем redirect
        pass
    return resp

# Инициализация расширений
db.init_app(app)
cache = Cache(app)

# Настройка логирования
if not Path('logs').exists():
    Path('logs').mkdir(parents=True, exist_ok=True)
if not app.debug:
    file_handler = RotatingFileHandler(str(Path('logs') / 'app.log'), maxBytes=10000, backupCount=10)
    file_handler.setFormatter(logging.Formatter(
        '%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]'
    ))
    file_handler.setLevel(logging.INFO)
    app.logger.addHandler(file_handler)
    
    app.logger.setLevel(logging.INFO)
    app.logger.info('Flask App startup')

# Настройка логирования
if not Path('logs').exists():
    Path('logs').mkdir(parents=True, exist_ok=True)
handler = RotatingFileHandler(str(Path('logs') / 'app.log'), maxBytes=10000, backupCount=3)
handler.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)
app.logger.addHandler(handler)
app.logger.setLevel(logging.INFO)

# Длительность "постоянной" сессии (например 30 дней)
app.permanent_session_lifetime = timedelta(days=30)

# Функция для безопасного сохранения в базу данных
def safe_commit():
    try:
        db.session.commit()
        return True
    except SQLAlchemyError as e:
        db.session.rollback()
        app.logger.error(f'Database error: {str(e)}')
        flash('Произошла ошибка базы данных')
        return False

# Legacy manual implementation removed — using the SDK-based implementation defined later in this file.
# The older manual POST-based function produced inconsistent 404 responses because the
# endpoint/format used didn't match the current Generative Language / Vision API.
# See `analyze_image_with_gemini` (SDK) defined further down.

# Декоратор для проверки авторизации
from functools import wraps

def login_required(role=None):
    """
    Декоратор проверки авторизации.

    - Если role is None: достаточно быть залогиненным (любая роль).
    - Если role задан (строка или список/кортеж/множество) — роль в сессии должна быть в разрешённых.
    Перенаправляет на "admin_login" только когда ожидаемая роль содержит 'admin',
    в остальных случаях использует стандартный маршрут 'login'.
    """
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            # пользователь не залогинен
            if 'user_id' not in session:
                flash('Сначала войдите в систему.', 'error')
                # Используем общий маршрут входа для всех ролей
                return redirect(url_for('login'))

            # если ожидается конкретная роль — проверим соответствие
            if role:
                # нормализуем allowed_roles в список
                if isinstance(role, (list, tuple, set)):
                    allowed = set(role)
                else:
                    allowed = {role}

                if session.get('role') not in allowed:
                    flash('Доступ запрещён!', 'error')
                    # Перенаправляем на общий вход
                    return redirect(url_for('login'))

            return f(*args, **kwargs)
        return decorated_function
    return decorator


# --- file upload settings (restore) ---
BASE_DIR = Path(__file__).resolve().parent
UPLOAD_FOLDER = BASE_DIR / 'static' / 'uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

# Создаем папку для загрузок, если её нет
UPLOAD_FOLDER.mkdir(parents=True, exist_ok=True)

from PIL import Image

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def save_image(file, max_size=(800, 800)):
    """
    Сохраняет и оптимизирует изображение
    """
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        timestamp = int(time.time())
        filename = f"{timestamp}_{filename}"

        # Создаем директорию если не существует
        UPLOAD_FOLDER.mkdir(parents=True, exist_ok=True)

        # Сохраняем оригинал
        filepath = UPLOAD_FOLDER / filename
        file.save(str(filepath))

        try:
            # Оптимизируем размер
            with Image.open(str(filepath)) as img:
                img.thumbnail(max_size)
                img.save(str(filepath), optimize=True, quality=85)
            app.logger.info(f'Image saved and optimized: {filename}')
        except Exception as e:
            app.logger.error(f'Image optimization failed: {str(e)}')

        return filename
    return None





# Контекст-процессор: предоставляет флаги авторизации в шаблоны и очищает невалидную сессию
@app.context_processor
def inject_user():
    user_id = session.get('user_id')
    role = session.get('role')
    if not user_id:
        return dict(logged_in=False, current_role=None)
    try:
        if role == 'student':
            user = Student.query.get(user_id)
        elif role == 'parent':
            user = Parents.query.get(user_id)
        elif role == 'cook':
            user = Cook.query.get(user_id)
        elif role == 'admin':
            user = Admin.query.get(user_id)
        else:
            user = None
    except Exception:
        user = None

    if user is None:
        # Очистим сессию, если данные невалидны
        session.pop('user_id', None)
        session.pop('role', None)
        session.pop('calories', None)
        session.pop('protein', None)
        session.pop('fat', None)
        session.pop('carbs', None)
        session.pop('eaten', None)
        return dict(logged_in=False, current_role=None)

    return dict(logged_in=True, current_role=role)

from typing import Union, Optional, Dict, List, Tuple

def safe_float(value: Union[str, int, float, None]) -> float:
    """Преобразует строки с запятой в float"""
    try:
        if value is None:
            return 0.0
        return float(str(value).replace(",", "."))
    except (ValueError, TypeError):
        return 0.0

def get_session_user_id() -> int:
    """Безопасно получает user_id из сессии"""
    try:
        user_id = session.get('user_id')
        if user_id is not None:
            return int(user_id)
        raise ValueError("No user_id in session")
    except (ValueError, TypeError):
        app.logger.error("Invalid user_id in session")
        raise ValueError("Invalid user_id in session")

def get_session_role() -> str:
    """Безопасно получает role из сессии"""
    role = session.get('role')
    if role not in ('student', 'parent', 'cook'):
        app.logger.error(f"Invalid role in session: {role}")
        raise ValueError("Invalid role in session")
    return role

def activity_to_coef(val: Union[str, float, None]) -> Optional[float]:
    """Преобразует строковое или числовое значение активности в числовой коэффициент.
    Принимает либо число в виде строки/float, либо ключи: minimal, light, medium, high, very_high
    Также поддерживаются русские варианты: "минимальная", "лёгкая", "средняя", "высокая", "очень высокая".
    Возвращает float или None, если не получилось распарсить.
    """
    if val is None or val == '':
        return None
    try:
        return float(val)
    except Exception:
        pass
    try:
        v = str(val).strip().lower()
    except Exception:
        return None
    mapping = {
        'minimal': 1.2, 'минимальная': 1.2, 'минимум': 1.2,
        'light': 1.375, 'лёгкая': 1.375, 'легкая': 1.375,
        'medium': 1.55, 'средняя': 1.55,
        'high': 1.725, 'высокая': 1.725,
        'very_high': 1.9, 'очень_высокая': 1.9, 'очень высокая': 1.9, 'оченьвысокая': 1.9
    }
    return mapping.get(v)

def get_nutrition_summary(logs: List[EatLog]) -> Dict[str, float]:
    """Считает суммарные показатели питания из логов"""
    sum_cal = sum(safe_float(log.calories) for log in logs)
    sum_prot = sum(safe_float(log.protein) for log in logs)
    sum_fat = sum(safe_float(log.fat) for log in logs)
    sum_carbs = sum(safe_float(log.carbs) for log in logs)
    
    return {
        'calories': round(sum_cal, 1),
        'protein': round(sum_prot, 1),
        'fat': round(sum_fat, 1),
        'carbs': round(sum_carbs, 1)
    }

def calculate_remaining_nutrients(total: Dict[str, float], consumed: Dict[str, float]) -> Dict[str, float]:
    """Считает оставшиеся показатели питания"""
    return {
        'calories': round(total['calories'] - consumed['calories'], 1),
        'protein': round(total['protein'] - consumed['protein'], 1),
        'fat': round(total['fat'] - consumed['fat'], 1),
        'carbs': round(total['carbs'] - consumed['carbs'], 1)
    }


def _extract_text_from_chunk(chunk):
    """Try to extract text from a chunk which can be str, dict, bytes, or have nested structures."""
    try:
        if chunk is None:
            return ''
        if isinstance(chunk, bytes):
            return chunk.decode('utf-8', errors='ignore')
        if isinstance(chunk, str):
            return chunk
        if isinstance(chunk, dict):
            # OpenAI-like: {'choices':[{'message':{'content': '...'}}]}
            if 'choices' in chunk and isinstance(chunk['choices'], (list, tuple)):
                parts = []
                for ch in chunk['choices']:
                    if isinstance(ch, dict):
                        if 'message' in ch and isinstance(ch['message'], dict) and ch['message'].get('content'):
                            parts.append(str(ch['message'].get('content')))
                        elif ch.get('text'):
                            parts.append(str(ch.get('text')))
                        elif ch.get('delta') and isinstance(ch.get('delta'), dict) and ch['delta'].get('content'):
                            parts.append(str(ch['delta'].get('content')))
                return ''.join(parts)
            # direct content
            if chunk.get('content'):
                return str(chunk.get('content'))
            if chunk.get('text'):
                return str(chunk.get('text'))
            # fallback: join values
            try:
                return ' '.join([str(v) for v in chunk.values()])
            except Exception:
                return str(chunk)
        # iterable (generator/list) -> join recursive
        if hasattr(chunk, '__iter__'):
            try:
                parts = []
                for c in chunk:
                    parts.append(_extract_text_from_chunk(c))
                return ''.join(parts)
            except Exception:
                return str(chunk)
    except Exception:
        return str(chunk)


def normalize_response(obj):
    """Normalize various g4f responses into a single text string."""
    return _extract_text_from_chunk(obj) or ''


def extract_clean_answer(text: str) -> str:
    """Try to strip metadata/dict dumps and return the final assistant message.

    Strategy:
    - If there's a trailing '}' (JSON-like), take substring after the last '}' and strip.
    - Otherwise, try to find the last block of Cyrillic or printable chars and return that.
    - Fallback to original text.
    """
    if not text:
        return ''
    try:
        import re
        import json

        # 1) If the output contains a JSON block(s), try to parse the last JSON object
        #    and extract common fields like choices[].message.content, content, text.
        last_json = None
        # try to find last opening brace and parse from there moving left
        for start in range(len(text) - 1, -1, -1):
            if text[start] == '{':
                candidate = text[start:]
                try:
                    parsed = json.loads(candidate)
                    last_json = parsed
                    break
                except Exception:
                    # not valid JSON from this position, continue
                    continue

        if last_json is not None:
            # try to extract assistant text from known shapes
            def extract_from_obj(obj):
                if obj is None:
                    return ''
                if isinstance(obj, str):
                    return obj
                if isinstance(obj, dict):
                    # OpenAI-like
                    if 'choices' in obj and isinstance(obj['choices'], (list, tuple)):
                        parts = []
                        for ch in obj['choices']:
                            if isinstance(ch, dict):
                                # g4f/OpenAI streaming shape
                                if ch.get('message') and isinstance(ch.get('message'), dict) and ch['message'].get('content'):
                                    parts.append(str(ch['message']['content']))
                                elif ch.get('text'):
                                    parts.append(str(ch['text']))
                                elif ch.get('delta') and isinstance(ch['delta'], dict) and ch['delta'].get('content'):
                                    parts.append(str(ch['delta'].get('content')))
                        if parts:
                            return ' '.join(parts)
                    # common keys
                    for key in ('assistant', 'response', 'content', 'text'):
                        val = obj.get(key)
                        if val:
                            return extract_from_obj(val)
                    # fallback: try to join string values
                    try:
                        return ' '.join([str(v) for v in obj.values() if isinstance(v, (str, int, float))])
                    except Exception:
                        return ''
                if isinstance(obj, (list, tuple)):
                    return ' '.join([extract_from_obj(x) for x in obj])
                return str(obj)

            extracted = extract_from_obj(last_json)
            if extracted and any(ch.isalpha() for ch in extracted):
                return extracted.strip()

        # 2) If JSON parsing didn't help: look for common assistant prefixes and return everything after them
        #    e.g. 'assistant:', 'Assistant:', 'Ответ:', 'Response:'
        prefixes = [r'assistant[:\-\s]+', r'Assistant[:\-\s]+', r'Ответ[:\-\s]+', r'Response[:\-\s]+', r'Ответ бота[:\-\s]+']
        for p in prefixes:
            m = re.search(p + r'(.+)$', text, re.IGNORECASE | re.DOTALL)
            if m:
                cand = m.group(1).strip()
                # drop any leading JSON-like garbage
                cand = re.sub(r'^\s*\{.*?\}\s*', '', cand, flags=re.DOTALL).strip()
                if cand:
                    return cand

        # 3) Heuristic: scan lines from bottom up and pick the first contiguous block that looks like natural language
        lines = text.splitlines()
        natural = []
        for line in reversed(lines):
            s = line.strip()
            if not s:
                if natural:
                    break
                else:
                    continue
            # skip obvious metadata lines
            if s.startswith('{') or s.startswith('[') or s.startswith('Traceback') or 'status_code' in s or s.lower().startswith('error'):
                if natural:
                    break
                else:
                    continue
            # skip short lines that look like JSON keys or file paths
            if len(s) < 20 and (':' in s and not re.search(r'[А-Яа-яA-Za-z]', s)):
                if natural:
                    break
                else:
                    continue
            natural.append(s)
            # collect a reasonable amount
            if sum(len(x) for x in natural) > 2000:
                break

        if natural:
            natural.reverse()
            return '\n'.join(natural).strip()

        # 4) final fallback: take last chunk of text and strip leading non-letter characters
        tail = text[-1200:]
        m = re.search(r'[\w\p{L}].*', tail)
        if m:
            return m.group(0).strip()
        return tail.strip()
    except Exception:
        return text.strip()


# ---------------- Регистрация ----------------

@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        try:
            role = request.form.get("role")
            login = request.form.get("login")  # Изменено с logine на login
            password = request.form.get("password")
            if not login or not password:  # Изменено с logine на login
                flash("Пожалуйста, заполните все поля")
                return redirect("/register")

            hash_pass = generate_password_hash(password)

            if role == "parent":
                # Родитель больше не указывает имя/метку ребёнка при регистрации.
                user = Parents(login=login, password=hash_pass)  # Изменено с logine на login

            elif role == "cook":
                city = request.form.get("city")
                school = request.form.get("school")
                user = Cook(login=login, password=hash_pass, city=city, school=school)  # Изменено с logine на login

            elif role == "teacher":
                # Создаем учителя как студента с особыми правами
                city = request.form.get("city")
                school = request.form.get("school")

                # Проверяем, что указаны обязательные поля
                if not city or not school:
                    flash("Пожалуйста, укажите город и школу")
                    return redirect("/register")

                # Попробуем прочитать антропометрические данные для расчёта КБЖУ
                gender = request.form.get('gender', 'male')
                measurements = {}
                for key in ('age', 'height', 'weight', 'activity'):
                    v = request.form.get(key)
                    try:
                        if v in (None, ''):
                            measurements[key] = None
                        elif key == 'activity':
                            measurements[key] = activity_to_coef(v)
                        else:
                            measurements[key] = float(v)
                    except Exception:
                        measurements[key] = None

                # Получаем значения КБЖУ из формы
                try:
                    cal = float(request.form.get('default_calories', 2000))
                    prot = float(request.form.get('default_protein', 75))
                    fat = float(request.form.get('default_fat', 60))
                    carbs = float(request.form.get('default_carbs', 250))
                except (TypeError, ValueError):
                    cal = 2000.0
                    prot = 75.0
                    fat = 60.0
                    carbs = 250.0

                # Если все измерения предоставлены — валидируем и считаем
                if all(measurements.get(k) is not None for k in ('age', 'height', 'weight', 'activity')):
                    try:
                        is_valid, err = validate_measurements(measurements)
                        if is_valid:
                            cal, prot, fat, carbs = calculate_nutrition(gender, measurements)
                        else:
                            flash(f"Ошибка измерений: {err}", 'warning')
                    except Exception as e:
                        app.logger.exception(f'Failed to calculate nutrition for teacher registration: {e}')

                # Создаём запись Student и задаём рассчитанные КБЖУ
                user = Student(
                    login=login,
                    password=hash_pass,
                    calories=cal,
                    protein=prot,
                    fat=fat,
                    carbs=carbs,
                    role="student",
                    city=city,
                    school=school
                )

            else:
                flash("Регистрация учеников через форму запрещена. Родитель должен добавить ребёнка в своём аккаунте.")
                return redirect("/register")

            # end of role handling
            try:
                db.session.add(user)
                db.session.commit()
                flash("Регистрация прошла успешно!")
                return redirect("/login")
            except Exception:
                db.session.rollback()
                app.logger.exception('Error committing new user in register')
                flash("Ошибка: логин уже существует или другая ошибка.")
                return redirect('/register')

        except Exception:
            app.logger.exception('Unexpected error in register POST')
            flash('Ошибка при регистрации (подробнее в логах).')
            return redirect('/register')

    return render_template("register.html")


# ---------------- Вход ----------------

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        login = request.form['login']
        password = request.form['password']

        # Проверяем администратора
        success, message = verify_admin(login, password)
        if success:
            admin = Admin.query.filter_by(login=login).first()
            session['user_id'] = admin.id
            session['role'] = 'admin'
            session.permanent = True  # Сессия будет жить до выхода
            return redirect(url_for('dashboard'))
        elif "не активирован" in message:
            flash('Аккаунт администратора не активирован.', 'error')
            return redirect(url_for('admin_activation'))

        # Проверяем сначала студента/учителя
        user = Student.query.filter_by(login=login).first()  # Изменено с logine на login
        if user and check_password_hash(user.password, password):
            session['user_id'] = user.id
            session['role'] = 'student'
            session["calories"] = user.calories
            session["protein"] = user.protein
            session["fat"] = user.fat
            session["carbs"] = user.carbs
            session["eaten"] = []
            session.permanent = True
            flash('Добро пожаловать, ученик!')
            return redirect(url_for('dashboard'))

        # Проверяем родителя
        user = Parents.query.filter_by(login=login).first()  # Изменено с logine на login
        if user and check_password_hash(user.password, password):
            session['user_id'] = user.id
            session['role'] = 'parent'
            session.permanent = True
            flash('Добро пожаловать, родитель!')
            return redirect(url_for('dashboard'))

    # Проверяем Диетсестра
        user = Cook.query.filter_by(login=login).first()  # Изменено с logine на login
        if user and check_password_hash(user.password, password):
            session['user_id'] = user.id
            session['role'] = 'cook'
            session.permanent = True
            flash(f"Добро пожаловать, {get_role_display('cook')}!")
            return redirect(url_for('dashboard'))

        flash('Неверный логин или пароль.', 'error')
        return redirect(url_for('login'))

    return render_template('login.html')


# ---------------- Dashboard ----------------

@app.route('/dashboard')
def dashboard():
    if 'user_id' not in session:
        flash('Сначала войдите в систему.')
        return redirect(url_for('login'))

    role = session.get('role')
    user_id = session.get('user_id')
    template_data = {}

    if role == 'student':
        user = Student.query.get(user_id)
        if not user:
            session.clear()
            flash('Ошибка: пользователь не найден.')
            return redirect(url_for('login'))

        # Получаем логи за сегодня из БД, чтобы состояние было едино для любых устройств
        start_of_day = datetime.combine(date.today(), datetime.min.time())
        today_logs = EatLog.query.filter(
            db.and_(EatLog.student_id == user.id, EatLog.created_at >= start_of_day)
        ).order_by(EatLog.created_at.asc()).all()

        sum_cal = sum([safe_float(l.calories) for l in today_logs])
        sum_prot = sum([safe_float(l.protein) for l in today_logs])
        sum_fat = sum([safe_float(l.fat) for l in today_logs])
        sum_carbs = sum([safe_float(l.carbs) for l in today_logs])

        remaining_cal = round(safe_float(user.calories) - sum_cal, 1)
        remaining_prot = round(safe_float(user.protein) - sum_prot, 1)
        remaining_fat = round(safe_float(user.fat) - sum_fat, 1)
        remaining_carbs = round(safe_float(user.carbs) - sum_carbs, 1)

        template_data = {
            'role': 'student',
            'login': user.login,
            'name': user.name,
            'calories': remaining_cal,
            'protein': remaining_prot,
            'fat': remaining_fat,
            'carbs': remaining_carbs,
            'eaten': today_logs,
            'view_as_student': True
        }

    elif role == 'parent':
        user = Parents.query.get(user_id)
        if not user:
            session.clear()
            flash('Ошибка: пользователь не найден.')
            return redirect(url_for('login'))
        
        children = Student.query.filter_by(parent_id=user.id).all()
        start_of_day = datetime.combine(date.today(), datetime.min.time())
        child_summaries = {}
        child_logs = {}
        for c in children:
            logs = EatLog.query.filter(EatLog.student_id == c.id, EatLog.created_at >= start_of_day).all()
            child_logs[c.id] = logs
            sum_cal = sum([safe_float(l.calories) for l in logs])
            sum_prot = sum([safe_float(l.protein) for l in logs])
            sum_fat = sum([safe_float(l.fat) for l in logs])
            sum_carbs = sum([safe_float(l.carbs) for l in logs])
            remaining_cal = round(safe_float(c.calories) - sum_cal, 1)
            remaining_prot = round(safe_float(c.protein) - sum_prot, 1)
            remaining_fat = round(safe_float(c.fat) - sum_fat, 1)
            remaining_carbs = round(safe_float(c.carbs) - sum_carbs, 1)
            child_summaries[c.id] = {
                'sum_cal': round(sum_cal, 1), 'sum_prot': round(sum_prot, 1), 
                'sum_fat': round(sum_fat, 1), 'sum_carbs': round(sum_carbs, 1),
                'remaining_cal': remaining_cal, 'remaining_prot': remaining_prot, 
                'remaining_fat': remaining_fat, 'remaining_carbs': remaining_carbs
            }
        
        template_data = {
            'role': 'parent',
            'login': user.login,
            'babe': user.babe,
            'children': children,
            'child_logs': child_logs,
            'child_summaries': child_summaries
        }

    elif role == 'cook':
        user = Cook.query.get(user_id)
        if not user:
            session.clear()
            flash('Ошибка: пользователь не найден.')
            return redirect(url_for('login'))
        template_data = {
            'role': 'cook',
            'login': user.login
        }


    else:
        flash('Неизвестная роль.')
        return redirect(url_for('login'))

    return render_template('dashboard.html', **template_data)


# ---------------- Выход ----------------

@app.route('/logout')
def logout():
    session.clear()
    flash('Вы вышли из системы.')
    return redirect(url_for('login'))  # login для всех ролей


# ---------------- Обновление КБЖУ вручную ----------------

@app.route('/update_calories', methods=['POST'])
def update_calories():
    if "user_id" not in session:
        return redirect(url_for('login'))

    user = Student.query.get(session["user_id"])
    if user:
        calories = request.form.get("calories")
        protein = request.form.get("protein")
        fat = request.form.get("fat")
        carbs = request.form.get("carbs")

        if calories and calories.isdigit():
            session["calories"] = int(calories)
        if protein:
            session["protein"] = float(protein)
        if fat:
            session["fat"] = float(fat)
        if carbs:
            session["carbs"] = float(carbs)

        session.modified = True
        flash("Данные успешно обновлены!", "success")

    return redirect(url_for('dashboard'))


# ---------------- Страница выбора еды ----------------

@app.route('/eat', methods=['GET', 'POST'])
@login_required()
def eat():
    # Ученики и учителя могут использовать выбор еды
    if session.get('role') not in ('student', 'teacher'):
        flash('Доступ запрещён', 'error')
        return redirect(url_for('login'))

    # Обработка поискового запроса (GET-параметр q)
    q = request.args.get('q', '').strip()
    barcode_q = request.args.get('barcode', '').strip()

    # Если задан штрихкод — ищем точное совпадение по barcode
    # Вычисляем текущую неделю/день для фильтрации школьного меню
    today = date.today()
    reference_date = date(2025, 1, 1)
    weeks_since = (today - reference_date).days // 7
    week = (weeks_since % 2) + 1
    weekday = today.weekday()  # 0=Mon .. 6=Sun
    day = weekday + 1

    if barcode_q:
        foods = Eat.query.filter_by(barcode=barcode_q).all()
        school_foods = [f for f in foods if f.type == 'school']
        other_foods = [f for f in foods if f.type != 'school']
    elif q:
        # Используем case-insensitive поиск через LIKE для SQLite
        search_pattern = f"%{q}%"
        foods = Eat.query.filter(
            db.or_(
                Eat.name.ilike(search_pattern),
                Eat.type.ilike(search_pattern)
            )
        ).all()
        school_foods = [f for f in foods if f.type == 'school' and getattr(f, 'week', None) == week and getattr(f, 'day', None) == day]
        other_foods = [f for f in foods if f.type != 'school']
    else:
        # По умолчанию показываем школьную еду только для текущей недели/дня и все прочие продукты
        school_foods = Eat.query.filter_by(type='school', week=week, day=day).all()
        other_foods = Eat.query.filter(Eat.type != 'school').all()

    # Объединяем найденные наборы в единый список для шаблона.
    # Это гарантирует, что переменная `foods` всегда определена (даже если пустая).
    try:
        foods = list(school_foods) + list(other_foods)
    except Exception:
        # на случай, если переменные не определены по какой-то причине — обеспечим безопасный fallback
        foods = []

    if request.method == 'POST':
        try:
            food_id = int(request.form.get('food_id') or 0)
        except ValueError:
            flash('Некорректный ID блюда')
            return redirect(url_for('eat'))

        grams = request.form.get('grams')
        food = Eat.query.get(food_id)

        if food:
            if "eaten" not in session:
                session["eaten"] = []


            if food.type == 'school':
                session["calories"] = round(session.get("calories", 0) - safe_float(food.calories), 1)
                session["protein"] = round(session.get("protein", 0) - safe_float(food.protein), 1)
                session["fat"] = round(session.get("fat", 0) - safe_float(food.fat), 1)
                session["carbs"] = round(session.get("carbs", 0) - safe_float(food.carbs), 1)
                session["eaten"].append({
                    'name': f"{food.name} (фикс)",
                    'calories': round(safe_float(food.calories), 1),
                    'protein': round(safe_float(food.protein), 1),
                    'fat': round(safe_float(food.fat), 1),
                    'carbs': round(safe_float(food.carbs), 1)
                })

                # логируем в базу
                try:
                    try:
                        student_id = get_session_user_id()
                        log = EatLog(
                            student_id=student_id,
                            food_id=food.id,
                            name=food.name,
                            calories=round(safe_float(food.calories), 1),
                            protein=round(safe_float(food.protein), 1),
                            fat=round(safe_float(food.fat), 1),
                            carbs=round(safe_float(food.carbs), 1)
                        )
                        db.session.add(log)
                        if not safe_commit():
                            flash('Ошибка при сохранении', 'error')
                    except ValueError:
                        flash('Ошибка: некорректный ID пользователя', 'error')
                        return redirect(url_for('login'))
                except Exception:
                    db.session.rollback()

            else:  # обычная еда
                try:
                    grams = float(grams or 0)  # масса блюда в граммах
                except ValueError:
                    grams = 0
                calc_cal = round(safe_float(food.calories) * grams / 100, 1)
                calc_prot = round(safe_float(food.protein) * grams / 100, 1)
                calc_fat = round(safe_float(food.fat) * grams / 100, 1)
                calc_carbs = round(safe_float(food.carbs) * grams / 100, 1)
                session["calories"] = round(session.get("calories", 0) - calc_cal, 1)
                session["protein"] = round(session.get("protein", 0) - calc_prot, 1)
                session["fat"] = round(session.get("fat", 0) - calc_fat, 1)
                session["carbs"] = round(session.get("carbs", 0) - calc_carbs, 1)
                session["eaten"].append({
                    'name': f"{food.name} ({grams} г)",
                    'calories': calc_cal,
                    'protein': calc_prot,
                    'fat': calc_fat,
                    'carbs': calc_carbs
                })

                # логируем в базу для обычной еды
                try:
                    factor = grams / 100.0
                    log = EatLog(student_id=session.get('user_id'), food_id=food.id, name=food.name,
                                 calories=round(safe_float(food.calories) * factor, 1),
                                 protein=round(safe_float(food.protein) * factor, 1),
                                 fat=round(safe_float(food.fat) * factor, 1),
                                 carbs=round(safe_float(food.carbs) * factor, 1))
                    db.session.add(log)
                    db.session.commit()
                except Exception:
                    db.session.rollback()

            session.modified = True
            flash(f"Вы добавили {food.name}!")

        return redirect(url_for('eat'))

    return render_template('eat.html', foods=foods, school_foods=school_foods, other_foods=other_foods, week=week, day=day)


# ---------------- Главная ----------------

@app.route('/favicon.ico')
def favicon():
    return send_from_directory('static/img', 'favicon.ico')

# Установка безопасных заголовков для всех ответов
@app.after_request
def set_security_headers(response):
    """Добавляет заголовки безопасности ко всем ответам."""
    try:
        if response:
            response.headers['X-Content-Type-Options'] = 'nosniff'
            response.headers['X-Frame-Options'] = 'SAMEORIGIN'
            response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
            response.headers['Permissions-Policy'] = 'camera=(), microphone=(), geolocation=()'
    except Exception as e:
        app.logger.warning(f'Failed to set security headers: {e}')
    return response

@app.route('/')
def index():
    """Показать школьное меню и кнопку анализа фото"""
    # Отладочный вывод для проверки языка и переводов
    app.logger.debug(f"Current language: {g.lang if hasattr(g, 'lang') else 'not set'}")
    app.logger.debug(f"Translations loaded: {bool(g.translations) if hasattr(g, 'translations') else False}")
    if hasattr(g, 'translations'):
        app.logger.debug(f"Sample translations: todays_menu='{g.translations.get('todays_menu', 'missing')}'")
    
    # Показать школьное меню для сегодняшнего рабочего дня по двухнедельному циклу
    today = date.today()
    weekday = today.weekday()  # 0=Mon .. 6=Sun

    # определяем двунедельный цикл: рассчитываем количество недель от опорной даты
    # Принято: цикл начинается от 2025-01-01. Если нужно другое начало — поменяйте reference_date.
    reference_date = date(2025, 1, 1)
    weeks_since = (today - reference_date).days // 7
    week = (weeks_since % 2) + 1
    day = weekday + 1

    foods_today = []
    if weekday <= 6:  # include Sunday for testing (0=Mon .. 6=Sun)
        foods_today = Eat.query.filter_by(type='school', week=week, day=day).all()

    # Передаём текущую неделю и день в шаблон, чтобы показывать пользователю
    return render_template('index.html', foods_today=foods_today, week=week, day=day, today=today)

# ---------------- Управление школьной едой для Диетсестра ----------------
def ensure_packs_exist(cook_id):
    """Создает 14 паков (2 недели × 7 дней), если они еще не существуют"""
    for week in (1, 2):
        for day in range(1, 8):  # 1-7
            pack = Pack.query.filter_by(week=week, day=day).first()
            if not pack:
                pack = Pack(week=week, day=day, created_by=cook_id)
                db.session.add(pack)
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        app.logger.exception('Failed to create packs')

@app.route('/cook_menu', methods=['GET', 'POST'])
def cook_menu():
    if 'user_id' not in session or session.get('role') != 'cook':
        flash("Доступ запрещён!")
        return redirect(url_for('login'))

    # Убедимся, что паки для всех дней существуют
    ensure_packs_exist(session.get('user_id'))

    foods = Eat.query.filter_by(type='school').order_by(Eat.week, Eat.day).all()
    packs = Pack.query.order_by(Pack.week, Pack.day).all()

    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'toggle_pack_item':
            try:
                item_id = int(request.form.get('item_id'))
                item = PackItem.query.get_or_404(item_id)
                item.is_active = not item.is_active
                db.session.commit()
                flash('Статус элемента обновлен', 'success')
            except Exception:
                db.session.rollback()
                flash('Ошибка при обновлении статуса', 'error')
            return redirect(url_for('cook_menu'))

        if action == 'remove_pack_item':
            try:
                item_id = int(request.form.get('item_id'))
                item = PackItem.query.get_or_404(item_id)
                db.session.delete(item)
                db.session.commit()
                flash('Элемент удален из пака', 'success')
            except Exception:
                db.session.rollback()
                flash('Ошибка при удалении элемента', 'error')
            return redirect(url_for('cook_menu'))
            
        # === Управление продуктами ===
        if action == 'add':
            name = request.form.get('name')
            calories = safe_float(request.form.get('calories'))
            protein = safe_float(request.form.get('protein'))
            fat = safe_float(request.form.get('fat'))
            carbs = safe_float(request.form.get('carbs'))
            
            try:
                week = int(request.form.get('week') or 0)
                if week not in (1, 2):
                    week = None
            except Exception:
                week = None

            try:
                day = int(request.form.get('day') or 0)
                if not (1 <= day <= 7):
                    day = None
            except Exception:
                day = None

            # Обработка файла изображения
            file = request.files.get('image_file')
            filename = None
            if file and file.filename and allowed_file(file.filename):
                fname = secure_filename(file.filename)
                import time
                fname = f"{int(time.time())}_{fname}"
                UPLOAD_FOLDER.mkdir(parents=True, exist_ok=True)
                file.save(str(UPLOAD_FOLDER / fname))
                filename = fname

            try:
                new_food = Eat(name=name, calories=calories, protein=protein,
                             fat=fat, carbs=carbs, type='school', 
                             image=filename, week=week, day=day)
                db.session.add(new_food)
                db.session.flush()  # чтобы получить id

                # Если указаны неделя и день - автоматически добавляем в соответствующий пак
                if week and day:
                    pack = Pack.query.filter_by(week=week, day=day).first()
                    if not pack:
                        pack = Pack(week=week, day=day, created_by=session.get('user_id'))
                        db.session.add(pack)
                        db.session.flush()
                    
                    # Находим максимальный порядковый номер
                    max_ord = db.session.query(db.func.max(PackItem.ord)).filter_by(pack_id=pack.id).scalar() or 0
                    
                    # Добавляем блюдо в пак
                    pack_item = PackItem(pack_id=pack.id, food_id=new_food.id, ord=max_ord + 1)
                    db.session.add(pack_item)

                db.session.commit()
                flash(f"Продукт {name} добавлен!")
            except Exception as e:
                db.session.rollback()
                app.logger.exception('Failed to add food')
                flash('Ошибка при добавлении продукта', 'error')

        elif action == 'edit':
            try:
                food_id = int(request.form.get('food_id') or 0)
            except ValueError:
                flash('Некорректный ID блюда')
                return redirect(url_for('cook_menu'))
                
            food = Eat.query.get(food_id)
            if food:
                old_week = food.week
                old_day = food.day
                
                food.name = request.form.get('name')
                food.calories = safe_float(request.form.get('calories'))
                food.protein = safe_float(request.form.get('protein'))
                food.fat = safe_float(request.form.get('fat'))
                food.carbs = safe_float(request.form.get('carbs'))

                try:
                    week = int(request.form.get('week') or 0)
                    if week not in (1, 2):
                        week = None
                except Exception:
                    week = None

                try:
                    day = int(request.form.get('day') or 0)
                    if not (1 <= day <= 7):
                        day = None
                except Exception:
                    day = None

                food.week = week
                food.day = day

                # Обработка файла изображения
                file = request.files.get('image_file')
                if file and file.filename and allowed_file(file.filename):
                    fname = secure_filename(file.filename)
                    import time
                    fname = f"{int(time.time())}_{fname}"
                    UPLOAD_FOLDER.mkdir(parents=True, exist_ok=True)
                    file.save(str(UPLOAD_FOLDER / fname))
                    food.image = fname

                try:
                    # Если изменились неделя/день - обновляем привязку к паку
                    if (old_week != week or old_day != day) and week and day:
                        # Удаляем из старого пака
                        PackItem.query.filter_by(food_id=food.id).delete()
                        
                        # Добавляем в новый пак
                        pack = Pack.query.filter_by(week=week, day=day).first()
                        if not pack:
                            pack = Pack(week=week, day=day, created_by=session.get('user_id'))
                            db.session.add(pack)
                            db.session.flush()
                        
                        # Находим максимальный порядковый номер
                        max_ord = db.session.query(db.func.max(PackItem.ord)).filter_by(pack_id=pack.id).scalar() or 0
                        
                        # Добавляем блюдо в пак
                        pack_item = PackItem(pack_id=pack.id, food_id=food.id, ord=max_ord + 1)
                        db.session.add(pack_item)

                    db.session.commit()
                    flash(f"Продукт {food.name} обновлён!")
                except Exception:
                    db.session.rollback()
                    flash('Ошибка при обновлении продукта', 'error')

        elif action == 'delete':
            try:
                food_id = int(request.form.get('food_id') or 0)
            except ValueError:
                flash('Некорректный ID блюда')
                return redirect(url_for('cook_menu'))
            
            food = Eat.query.get(food_id)
            if food:
                try:
                    # Сначала удаляем все связи с паками
                    PackItem.query.filter_by(food_id=food.id).delete()
                    db.session.delete(food)
                    db.session.commit()
                    flash(f"Продукт {food.name} удалён!")
                except Exception:
                    db.session.rollback()
                    flash('Ошибка при удалении продукта', 'error')

        return redirect(url_for('cook_menu'))

    return render_template('cook_menu.html', foods=foods, packs=packs)


@app.route('/pack_scan')
def pack_scan():
    """Страница просмотра пака по ссылке/QR (публичная)."""
    try:
        pack_id = request.args.get('pack_id', type=int)
        barcode = request.args.get('barcode')
        pack = None
        if pack_id:
            pack = Pack.query.get(pack_id)
        elif barcode:
            pack = Pack.query.filter_by(barcode=str(barcode).strip()).first()
        if not pack:
            flash('Пак не найден', 'error')
            return redirect(url_for('index'))

        items = PackItem.query.filter_by(pack_id=pack.id).order_by(PackItem.ord).all()
        logged_in = 'user_id' in session
        current_role = session.get('role')
        
        # Получаем информацию о съеденных блюдах для текущего пользователя
        eaten_today = {}
        if logged_in and current_role in ['student', 'parent']:
            try:
                # Определяем ID студента
                if current_role == 'student':
                    student_id = session.get('user_id')
                elif current_role == 'parent':
                    student_id = request.args.get('student_id', type=int)
                    # Проверяем, что родитель имеет доступ к этому ребенку
                    if student_id:
                        parent_id = session.get('user_id')
                        child = Student.query.get(student_id)
                        if not child or child.parent_id != parent_id:
                            student_id = None
                
                app.logger.info(f"Getting eaten foods for student_id: {student_id}, role: {current_role}")
                
                if student_id:
                    # Получаем логи питания за сегодня
                    today = datetime.utcnow().date()
                    today_logs = EatLog.query.filter(
                        EatLog.student_id == student_id,
                        func.date(EatLog.created_at) == today
                    ).all()
                    
                    app.logger.info(f"Found {len(today_logs)} logs for today")
                    
                    # Считаем количество каждого блюда
                    for log in today_logs:
                        if log.food_id not in eaten_today:
                            eaten_today[log.food_id] = 0
                        eaten_today[log.food_id] += 1
            except Exception as e:
                app.logger.exception('Error getting eaten foods')
        
        return render_template('pack_scan.html', 
                             pack=pack, 
                             items=items, 
                             logged_in=logged_in, 
                             current_role=current_role,
                             eaten_today=eaten_today)
    except Exception as e:
        app.logger.exception('Error in pack_scan')
        flash('Ошибка при отображении пака', 'error')
        return redirect(url_for('index'))


@app.route('/pack_add', methods=['POST'])
@login_required(role=['student', 'parent'])
def pack_add():
    """Добавить один продукт или весь пак в дневник пользователя (для student или parent).
    POST params: food_id or pack_id, optional student_id (when parent adds for child)
    """
    try:
        role = session.get('role')
        target_student_id = None
        if role == 'parent' and request.form.get('student_id'):
            try:
                target_student_id = int(request.form.get('student_id'))
            except Exception:
                target_student_id = None
            # verify ownership
            if target_student_id:
                parent_id = get_session_user_id()
                child = Student.query.get(target_student_id)
                if not child or child.parent_id != parent_id:
                    flash('Ребёнок не найден или доступ запрещён', 'error')
                    return redirect(url_for('parent_children'))

        def add_log_for_student(sid, food_obj, servings=1):
            try:
                # Умножаем все значения на количество порций
                servings = float(servings)
                calories = round(safe_float(food_obj.calories) * servings, 1)
                protein = round(safe_float(food_obj.protein) * servings, 1)
                fat = round(safe_float(food_obj.fat) * servings, 1)
                carbs = round(safe_float(food_obj.carbs) * servings, 1)
                log = EatLog(student_id=sid, food_id=food_obj.id, 
                           name=f"{food_obj.name} x{int(servings) if servings.is_integer() else servings}",
                           calories=calories, protein=protein, fat=fat, carbs=carbs)
                db.session.add(log)
                db.session.commit()
                return True
            except Exception:
                db.session.rollback()
                return False

        # single food
        food_id = request.form.get('food_id')
        if food_id:
            try:
                fid = int(food_id)
            except Exception:
                flash('Некорректный ID блюда', 'error')
                return redirect(url_for('index'))
            food = Eat.query.get(fid)
            if not food:
                flash('Блюдо не найдено', 'error')
                return redirect(url_for('index'))

            # Получаем количество порций
            try:
                servings = float(request.form.get('servings', 1))
                if servings < 0.1:  # Минимальное допустимое значение
                    servings = 1
            except (ValueError, TypeError):
                servings = 1

            if role == 'student' and not target_student_id:
                sid = get_session_user_id()
                ok = add_log_for_student(sid, food, servings)
                if ok:
                    # Обновляем значения в сессии (без знака минус, так как это еда которую съели)
                    student = Student.query.get(sid)
                    session['calories'] = round(safe_float(student.calories) - safe_float(food.calories) * servings, 1)
                    session['protein'] = round(safe_float(student.protein) - safe_float(food.protein) * servings, 1)
                    session['fat'] = round(safe_float(student.fat) - safe_float(food.fat) * servings, 1)
                    session['carbs'] = round(safe_float(student.carbs) - safe_float(food.carbs) * servings, 1)
                    
                    # Добавляем в список съеденного
                    eaten = session.get('eaten', [])
                    food_name = f"{food.name} x{int(servings) if servings.is_integer() else servings}"
                    eaten.append({
                        'name': food_name,
                        'calories': round(safe_float(food.calories) * servings, 1),
                        'protein': round(safe_float(food.protein) * servings, 1),
                        'fat': round(safe_float(food.fat) * servings, 1),
                        'carbs': round(safe_float(food.carbs) * servings, 1)
                    })
                    session['eaten'] = eaten
                    session.modified = True
                    
                    flash(f'Добавлено: {food_name}')
                else:
                    flash('Ошибка при добавлении', 'error')
                return redirect(url_for('dashboard'))
            else:
                # parent adding for child
                sid = target_student_id if target_student_id else get_session_user_id()
                ok = add_log_for_student(sid, food)
                if ok:
                    flash(f'Блюдо {food.name} добавлено в дневник')
                else:
                    flash('Ошибка при добавлении', 'error')
                return redirect(url_for('parent_children') if role == 'parent' else url_for('dashboard'))

        # pack add all
        pack_id = request.form.get('pack_id')
        if pack_id:
            try:
                pid = int(pack_id)
            except Exception:
                flash('Некорректный ID пака', 'error')
                return redirect(url_for('index'))
            pack = Pack.query.get(pid)
            if not pack:
                flash('Пак не найден', 'error')
                return redirect(url_for('index'))
            items = PackItem.query.filter_by(pack_id=pack.id).all()
            # determine target student id
            if role == 'student' and not target_student_id:
                sid = get_session_user_id()
            else:
                sid = target_student_id if target_student_id else get_session_user_id()
            added = 0
            student = Student.query.get(sid) if sid == session.get('user_id') else None
            eaten = session.get('eaten', []) if sid == session.get('user_id') else []
            
            for it in items:
                food = Eat.query.get(it.food_id)
                if food:
                    if add_log_for_student(sid, food, 1):  # добавляем по 1 порции
                        added += 1
                        # if current session belongs to that student, update session totals
                        if sid == session.get('user_id'):
                            # Добавляем в список съеденного
                            eaten.append({
                                'name': food.name,
                                'calories': round(safe_float(food.calories), 1),
                                'protein': round(safe_float(food.protein), 1),
                                'fat': round(safe_float(food.fat), 1),
                                'carbs': round(safe_float(food.carbs), 1)
                            })
            
            if sid == session.get('user_id') and student:
                # Обновляем суммарные значения в сессии
                total_cal = sum(item['calories'] for item in eaten)
                total_prot = sum(item['protein'] for item in eaten)
                total_fat = sum(item['fat'] for item in eaten)
                total_carbs = sum(item['carbs'] for item in eaten)
                
                session['calories'] = round(safe_float(student.calories) - total_cal, 1)
                session['protein'] = round(safe_float(student.protein) - total_prot, 1)
                session['fat'] = round(safe_float(student.fat) - total_fat, 1)
                session['carbs'] = round(safe_float(student.carbs) - total_carbs, 1)
                session['eaten'] = eaten
                session.modified = True
                
            flash(f'Добавлено {added} блюд из пака')
            return redirect(url_for('dashboard') if role == 'student' else url_for('parent_children'))

    except Exception as e:
        app.logger.exception('Error in pack_add')
        flash('Произошла ошибка при добавлении', 'error')
    return redirect(url_for('index'))


@app.route('/scan')
@login_required(role='student')
def scan():
    """
    Обработка сканирования QR-кода продукта.
    URL format: /scan?food_id=ID[&grams=NN] или /scan?barcode=CODE[&grams=NN]
    """
    # Получение и валидация параметров
    try:
        student_id = get_session_user_id()
        
        try:
            food_id = request.args.get('food_id', type=int)
        except ValueError:
            food_id = None
            
        barcode = request.args.get('barcode')
        if barcode:
            barcode = barcode.strip()
            
        try:
            grams = request.args.get('grams', type=float)
            if grams and (grams <= 0 or grams > 5000):  # разумное ограничение
                raise ValueError("Invalid grams value")
        except ValueError:
            grams = None

        # Проверка наличия идентификатора продукта
        if not food_id and not barcode:
            flash('Некорректный QR-код (нет id или штрихкода).', 'error')
            return redirect(url_for('dashboard'))
            
        # Поиск продукта
        if food_id:
            food = Eat.query.get(food_id)
        elif barcode:
            food = Eat.query.filter_by(barcode=barcode).first()
            
        if not food:
            flash('Продукт не найден', 'error')
            return redirect(url_for('dashboard'))
            
    except ValueError as e:
        app.logger.error(f'Error in scan: {str(e)}')
        flash('Ошибка при обработке запроса', 'error')
        return redirect(url_for('dashboard'))
        
    except Exception as e:
        app.logger.error(f'Unexpected error in scan: {str(e)}')
        flash('Произошла неожиданная ошибка', 'error')
        return redirect(url_for('dashboard'))

    food = None
    if food_id:
        try:
            food = Eat.query.get(int(food_id))
        except Exception:
            food = None
    elif barcode:
        try:
            food = Eat.query.filter_by(barcode=str(barcode).strip()).first()
        except Exception:
            food = None

    if not food:
        flash('Блюдо не найдено.')
        return redirect(url_for('dashboard'))

    if 'eaten' not in session:
        session['eaten'] = []

    # для школьной еды фиксированное снятие
    if food.type == 'school':
        session["calories"] = round(session.get("calories", 0) - safe_float(food.calories), 1)
        session["protein"] = round(session.get("protein", 0) - safe_float(food.protein), 1)
        session["fat"] = round(session.get("fat", 0) - safe_float(food.fat), 1)
        session["carbs"] = round(session.get("carbs", 0) - safe_float(food.carbs), 1)
        session['eaten'].append(f"{food.name} (фикс)")
        try:
            log = EatLog(student_id=session.get('user_id'), food_id=food.id, name=food.name,
                         calories=round(safe_float(food.calories), 1), protein=round(safe_float(food.protein), 1),
                         fat=round(safe_float(food.fat), 1), carbs=round(safe_float(food.carbs), 1))
            db.session.add(log)
            db.session.commit()
        except Exception:
            db.session.rollback()
    else:
        # если указан вес в QR, учитываем его
        if grams:
            g = safe_float(grams)
            session["calories"] = round(session.get("calories", 0) - safe_float(food.calories) * g / 100, 1)
            session["protein"] = round(session.get("protein", 0) - safe_float(food.protein) * g / 100, 1)
            session["fat"] = round(session.get("fat", 0) - safe_float(food.fat) * g / 100, 1)
            session["carbs"] = round(session.get("carbs", 0) - safe_float(food.carbs) * g / 100, 1)
            session['eaten'].append(f"{food.name} ({g} г)")
            try:
                factor = g / 100.0
                log = EatLog(student_id=session.get('user_id'), food_id=food.id, name=food.name,
                             calories=round(safe_float(food.calories) * factor, 1),
                             protein=round(safe_float(food.protein) * factor, 1),
                             fat=round(safe_float(food.fat) * factor, 1),
                             carbs=round(safe_float(food.carbs) * factor, 1))
                db.session.add(log)
                db.session.commit()
            except Exception:
                db.session.rollback()
        else:
            session["calories"] = round(session.get("calories", 0) - safe_float(food.calories), 1)
            session["protein"] = round(session.get("protein", 0) - safe_float(food.protein), 1)
            session["fat"] = round(session.get("fat", 0) - safe_float(food.fat), 1)
            session["carbs"] = round(session.get("carbs", 0) - safe_float(food.carbs), 1)
            session['eaten'].append(f"{food.name}")

    session.modified = True
    flash('Блюдо добавлено после сканирования QR-кода.')
    return redirect(url_for('dashboard'))


# Родитель может добавить ребёнка
@app.route('/parent/add_child', methods=['GET', 'POST'])
@login_required(role='parent')
def parent_add_child():
    from forms import ChildForm
    form = ChildForm()

    # Получаем списки для селекторов
    cities = City.query.order_by(City.name).all()
    schools = School.query.order_by(School.name).all()
    grades = Grade.query.order_by(Grade.name).all()

    # Динамически заполняем choices для WTForms
    form.city.choices = [('', '— выберите город —')] + [(c.name, c.name) for c in cities]
    form.school.choices = [('', '— выберите школу —')] + [(s.name, s.name) for s in schools]
    form.grade.choices = [('', '— выберите класс —')] + [(g.name, g.name) for g in grades]

    if form.validate_on_submit():
        try:
            parent_id = get_session_user_id()

            if not form.login.data or not form.password.data:
                flash('Логин и пароль обязательны')
                return redirect(url_for('parent_add_child'))

            # Подготовка измерений
            try:
                # Activity can be either numeric or a textual key (e.g. 'medium')
                raw_activity = form.activity.data if hasattr(form, 'activity') else None
                act_val = None
                if raw_activity not in (None, ''):
                    try:
                        act_val = float(raw_activity)
                    except Exception:
                        act_val = activity_to_coef(raw_activity)

                measurements = {
                    'age': float(form.age.data) if form.age.data is not None else None,
                    'height': float(form.height.data) if form.height.data is not None else None,
                    'weight': float(form.weight.data) if form.weight.data is not None else None,
                    'activity': act_val,
                }
            except (ValueError, TypeError):
                measurements = {'age': None, 'height': None, 'weight': None, 'activity': None}

            # Значения по умолчанию
            calories = 2000.0
            protein = round((calories * 0.2) / 4, 1)
            fat = round((calories * 0.3) / 9, 1)
            carbs = round((calories * 0.5) / 4, 1)

            # Если все данные предоставлены, валидируем и считаем
            if all(v is not None for v in measurements.values()):
                # приведение к Dict[str, float] для валидаторов
                measurements_float: dict[str, float] = {k: float(v) for k, v in measurements.items()}  # type: ignore
                is_valid, err = validate_measurements(measurements_float)
                if is_valid:
                    try:
                        calories, protein, fat, carbs = calculate_nutrition(gender=str(form.gender.data or 'male'), measurements=measurements_float)
                        # BMI
                        bmi = measurements_float['weight'] / ((measurements_float['height']/100.0)**2)
                        if bmi < 18.5 or calories < 1800:
                            flash('Рекомендуется увеличить суточную калорийность для ребёнка.', 'info')
                        elif bmi > 25 or calories > 2200:
                            flash('Рекомендуется уменьшить суточную калорийность для ребёнка.', 'info')
                    except Exception as e:
                        app.logger.warning(f'Nutrition calc failed: {e}')
                else:
                    app.logger.warning(f'Invalid measurements: {err}')

            # Создаем ученика
            try:
                hash_pass = generate_password_hash(str(form.password.data))
                # Определяем город/школу/класс — берём из select-полей формы
                city_val = None
                school_val = None
                grade_val = None
                try:
                    city_val = (form.city.data or None) if hasattr(form, 'city') else (request.form.get('city') or None)
                    school_val = (form.school.data or None) if hasattr(form, 'school') else (request.form.get('school') or None)
                    grade_val = (str(form.grade.data) if form.grade.data else None) if hasattr(form, 'grade') else (request.form.get('grade') or None)
                except Exception:
                    city_val = request.form.get('city') or None
                    school_val = request.form.get('school') or None
                    grade_val = request.form.get('grade') or None

                child = Student(
                    login=form.login.data,
                    password=hash_pass,
                    parent_id=parent_id,
                    calories=calories,
                    protein=protein,
                    fat=fat,
                    carbs=carbs,
                    city=city_val,
                    school=school_val,
                    grade=grade_val,
                    age=measurements.get('age') if isinstance(measurements, dict) else None,
                    height=measurements.get('height') if isinstance(measurements, dict) else None,
                    weight=measurements.get('weight') if isinstance(measurements, dict) else None,
                    activity=measurements.get('activity') if isinstance(measurements, dict) else None,
                    gender=(form.gender.data if hasattr(form, 'gender') else None)
                )
                db.session.add(child)
                if safe_commit():
                    app.logger.info(f'Child added: {child.login}')
                    flash('Ребёнок успешно добавлен!')
                    return redirect(url_for('dashboard'))
                else:
                    flash('Ошибка при сохранении', 'error')
            except Exception as e:
                app.logger.exception('Error adding child')
                flash('Ошибка при добавлении ребёнка (подробнее в логах).', 'error')

        except ValueError:
            flash('Ошибка сессии. Пожалуйста, войдите снова.', 'error')
            return redirect(url_for('login'))  # login для родителя
        except Exception:
            app.logger.exception('Unexpected error in parent_add_child')
            flash('Произошла неожиданная ошибка (подробнее в логах)', 'error')

    return render_template('parent_add_child.html', form=form, cities=cities, schools=schools, grades=grades)


# Родитель видит список своих детей и их логи
@app.route('/parent/children')
@login_required(role='parent')
# Кэшируем по-минуте, но включаем user_id в префикс ключа, чтобы кэш был пер-пользовательским
@cache.cached(timeout=60, key_prefix=lambda: f"parent_children:{session.get('user_id')}")
def parent_children():
    """Показывает список детей родителя и их логи питания за сегодня"""
    try:
        # Получаем и проверяем id родителя
        parent_id = get_session_user_id()
        parent = Parents.query.get_or_404(parent_id)
        
        # Используем join для оптимизации запроса
        children = (Student.query
                   .filter_by(parent_id=parent_id)
                   .order_by(Student.login)
                   .all())
                   
        if not children:
            flash('У вас пока нет добавленных детей')
            return redirect(url_for('parent_add_child'))
            
        # Инициализируем структуры для логов
        child_logs: Dict[int, List[EatLog]] = {}
        child_summaries: Dict[int, Dict[str, Dict[str, float]]] = {}
        start_of_day = datetime.combine(date.today(), datetime.min.time())
        
        # Получаем логи и считаем суммы для каждого ребенка
        for child in children:
            # Получаем логи за сегодня
            logs = EatLog.query.filter(
                db.and_(
                    EatLog.student_id == child.id,
                    EatLog.created_at >= start_of_day
                )
            ).all()
            child_logs[child.id] = logs
            
            # Используем тот же алгоритм, что и в dashboard, чтобы копировать вывод БЖУ
            # (суммы по логам и остатки = цель - сумма). Это гарантирует идентичный формат.
            sum_cal = sum([safe_float(l.calories) for l in logs])
            sum_prot = sum([safe_float(l.protein) for l in logs])
            sum_fat = sum([safe_float(l.fat) for l in logs])
            sum_carbs = sum([safe_float(l.carbs) for l in logs])

            remaining_cal = round(safe_float(child.calories) - sum_cal, 1)
            remaining_prot = round(safe_float(child.protein) - sum_prot, 1)
            remaining_fat = round(safe_float(child.fat) - sum_fat, 1)
            remaining_carbs = round(safe_float(child.carbs) - sum_carbs, 1)

            child_summaries[child.id] = {
                'sum_cal': round(sum_cal, 1),
                'sum_prot': round(sum_prot, 1),
                'sum_fat': round(sum_fat, 1),
                'sum_carbs': round(sum_carbs, 1),
                'remaining_cal': remaining_cal,
                'remaining_prot': remaining_prot,
                'remaining_fat': remaining_fat,
                'remaining_carbs': remaining_carbs
            }
            
            # Логируем для мониторинга (используем summary, чтобы не ссылаться на несуществующую переменную)
            try:
                app.logger.info(f"Calculated nutrition for child {child.id}: summary={child_summaries[child.id]}")
            except Exception:
                app.logger.info(f"Calculated nutrition for child {child.id}: sums computed")
            
    except ValueError as e:
        app.logger.error(f'Session error in parent_children: {str(e)}')
        flash('Ошибка сессии', 'error')
        return redirect(url_for('login'))
        
    except Exception as e:
        app.logger.error(f'Error in parent_children: {str(e)}')
        flash('Произошла ошибка', 'error')
        return redirect(url_for('dashboard'))
        
    # Diagnostic log: dump summary values so we can debug zero-values issue
    try:
        app.logger.info(f"parent_children: child_summaries dump: {child_summaries}")
    except Exception:
        app.logger.exception('Failed to log child_summaries in parent_children')

    return render_template('parent_children.html',
                         children=children,
                         child_logs=child_logs,
                         child_summaries=child_summaries)


@app.route('/parent/recalc_child/<int:student_id>', methods=['GET', 'POST'])
@login_required(role='parent')
def parent_recalc_child(student_id: int):
    """Пересчитать КБЖУ для ребёнка — доступно только родителю, у которого этот ребёнок привязан."""
    try:
        parent_id = get_session_user_id()
    except ValueError:
        flash('Ошибка сессии. Пожалуйста, войдите снова.', 'error')
        return redirect(url_for('login'))

    child = Student.query.get_or_404(student_id)
    if child.parent_id != parent_id:
        flash('Доступ запрещён', 'error')
        return redirect(url_for('parent_children'))

    # GET — показываем страницу с калькулятором/формой для ввода измерений
    if request.method == 'GET':
        return render_template('parent_recalc.html', child=child)

    # Попробуем получить новые значения из POST (если пользователь ввёл их в модале)
    form_age = request.form.get('age')
    form_height = request.form.get('height')
    form_weight = request.form.get('weight')
    form_activity = request.form.get('activity')
    form_gender = request.form.get('gender')

    # Если в форме переданы значения, используем их (и сохраним в модель). Иначе используем сохранённые значения.
    def parse_float_or_none(val):
        try:
            if val is None or val == '':
                return None
            return float(val)
        except Exception:
            return None

    age_val = parse_float_or_none(form_age) if form_age is not None else (float(child.age) if child.age is not None else None)
    height_val = parse_float_or_none(form_height) if form_height is not None else (float(child.height) if child.height is not None else None)
    weight_val = parse_float_or_none(form_weight) if form_weight is not None else (float(child.weight) if child.weight is not None else None)

    # activity может быть числовым или строковым (ключ). Попробуем распарсить
    activity_val = None
    if form_activity is not None and form_activity != '':
        try:
            activity_val = float(form_activity)
        except Exception:
            # map common keys to coef
            mapping = {'minimal':1.2,'light':1.375,'medium':1.55,'high':1.725,'very_high':1.9}
            activity_val = mapping.get(form_activity.lower(), None)
    else:
        if child.activity is not None:
            try:
                activity_val = float(child.activity)
            except Exception:
                activity_val = None

    gender_val = form_gender if (form_gender is not None and form_gender != '') else (child.gender or None)

    # Если пользователь ввёл хотя бы одно новое значение — сохраним их для ребёнка
    if form_age is not None or form_height is not None or form_weight is not None or form_activity is not None or form_gender is not None:
        if age_val is not None:
            child.age = age_val
        if height_val is not None:
            child.height = height_val
        if weight_val is not None:
            child.weight = weight_val
        if activity_val is not None:
            child.activity = activity_val
        if gender_val is not None:
            child.gender = gender_val
        try:
            db.session.add(child)
            db.session.commit()
        except Exception:
            db.session.rollback()

    # Если передан override калорий — сохраняем напрямую (и вычисляем макрораспределение)
    calories_override = request.form.get('calories_override')
    if calories_override:
        try:
            cal_val = float(calories_override)
            # Простое макро-распределение: 20% белки, 30% жиры, 50% углеводы
            prot = round((cal_val * 0.2) / 4, 1)
            fat = round((cal_val * 0.3) / 9, 1)
            carbs = round((cal_val * 0.5) / 4, 1)
            child.calories = float(cal_val)
            child.protein = float(prot)
            child.fat = float(fat)
            child.carbs = float(carbs)
            db.session.add(child)
            db.session.commit()
            try:
                cache.delete_memoized(parent_children)
            except Exception:
                app.logger.exception('Failed to delete cache for parent_children')
            flash('КБЖУ ребёнка сохранены (override).', 'success')
            return redirect(url_for('parent_children'))
        except Exception as e:
            app.logger.exception(f'Failed to save calories_override for child {child.id}: {e}')
            db.session.rollback()
            flash('Ошибка при сохранении переопределённых калорий.', 'error')
            return redirect(url_for('parent_children'))

    # Проверяем, что теперь есть все необходимые значения для расчёта
    if age_val is None or height_val is None or weight_val is None or activity_val is None:
        flash('Недостаточно данных для пересчёта КБЖУ. Укажите возраст, рост, вес и активность или используйте принудительное значение калорий.', 'warning')
        return redirect(url_for('parent_children'))

    measurements = {'age': float(age_val), 'height': float(height_val), 'weight': float(weight_val), 'activity': float(activity_val)}
    is_valid, err = validate_measurements(measurements)
    if not is_valid:
        flash(f'Невозможно пересчитать: {err}', 'error')
        return redirect(url_for('parent_children'))

    try:
        # Используем сохранённый пол ребёнка, если он есть
        gender = (child.gender or 'male')
        calories, protein, fat, carbs = calculate_nutrition(gender=gender, measurements=measurements)
        child.calories = float(calories)
        child.protein = float(protein)
        child.fat = float(fat)
        child.carbs = float(carbs)
        db.session.add(child)
        db.session.commit()
        # Убедимся, что кэш страницы списка детей обновится сразу
        try:
            cache.delete_memoized(parent_children)
        except Exception:
            app.logger.exception('Failed to delete cache for parent_children')
        flash('КБЖУ ребёнка пересчитаны и сохранены.', 'success')
    except Exception as e:
        app.logger.exception(f'Error recalculating nutrition for child {child.id}: {e}')
        db.session.rollback()
        flash('Ошибка при пересчёте или сохранении КБЖУ.', 'error')

    return redirect(url_for('parent_children'))


from services.google_sheets import create_nutrition_report


@app.route('/api/admin_login', methods=['POST'])
def admin_login_api():
    """API endpoint для входа админа через консоль браузера"""
    try:
        data = request.get_json()
        login = data.get('login')
        password = data.get('password')
        
        if not login or not password:
            return jsonify({'success': False, 'error': 'Требуется логин и пароль'})
            
        admin = Admin.query.filter_by(login=login).first()
        if admin and check_password_hash(admin.password, password):
            session['user_id'] = admin.id
            session['role'] = 'admin'
            session.permanent = True
            return jsonify({'success': True, 'redirect': url_for('dashboard')})
        else:
            return jsonify({'success': False, 'error': 'Неверный логин или пароль'})
            
    except Exception as e:
        app.logger.exception('Ошибка при входе админа через API')
        return jsonify({'success': False, 'error': str(e)})


@app.route('/adm/<login>/<password>')
def admin_quick_login(login, password):
    """Быстрый GET-роут для входа админа по URL: /adm/<login>/<password>
    При успешной аутентификации сразу редиректит на панель администратора.
    """
    try:
        admin = Admin.query.filter_by(login=login).first()
        if admin and check_password_hash(admin.password, password):
            session['user_id'] = admin.id
            session['role'] = 'admin'
            session.permanent = True
            app.logger.info(f'Quick admin login: {login}')
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Неверный логин или пароль', 'error')
            return redirect(url_for('login'))
    except Exception:
        app.logger.exception('Ошибка при быстром входе админа')
        flash('Ошибка при входе', 'error')
        return redirect(url_for('login'))

@app.route('/parent/child/<int:student_id>/export')
@login_required(role='parent')
def export_child_year(student_id: int):
    """Экспортирует логи питания ребёнка за последний год в Google Sheets.
    
    Параметры query: fmt=excel|word|sheets (по умолчанию sheets)
    Доступен только родителю, у которого этот ребёнок привязан.
    """
    fmt = request.args.get('fmt', 'sheets').lower()
    # Early-entry diagnostic log: confirm request reached export handler
    try:
        app.logger.info(
            f"EXPORT ENTRY: student_id={student_id} user_id={session.get('user_id')} role={session.get('role')} "
            f"remote={request.remote_addr} UA={request.headers.get('User-Agent')[:200]} fmt={fmt}"
        )
    except Exception:
        app.logger.exception('Failed to write EXPORT ENTRY log')
    # Проверка прав: родитель должен владеть ребёнком
    try:
        parent_id = get_session_user_id()
    except ValueError:
        flash('Ошибка сессии. Пожалуйста, войдите снова.', 'error')
        return redirect(url_for('login'))

    child = Student.query.get_or_404(student_id)
    if child.parent_id != parent_id:
        flash('Доступ запрещён', 'error')
        return redirect(url_for('parent_children'))

    # Получаем все логи, сортируем по дате
    logs = EatLog.query.filter(
        EatLog.student_id == child.id
    ).order_by(EatLog.created_at.asc()).all()
    
    # Находим первую и последнюю дату
    if logs:
        start = logs[0].created_at
        now = logs[-1].created_at
    else:
        now = datetime.utcnow()
        start = now

    # Целевые показатели (нормы) — делаем их доступными для всех форматов
    targets = {
        'calories': float(child.calories or 2000),
        'protein': float(child.protein or 75),
        'fat': float(child.fat or 60),
        'carbs': float(child.carbs or 250)
    }

    if fmt == 'sheets':
        # Преобразуем логи в формат для Google Sheets
        logs_data = [{
            'created_at': log.created_at,
            'name': log.name,
            'calories': float(log.calories or 0),
            'protein': float(log.protein or 0),
            'fat': float(log.fat or 0),
            'carbs': float(log.carbs or 0)
        } for log in logs]

        # Целевые показатели уже сформированы выше in `targets`

        # Создаем отчет
        sheet_url = create_nutrition_report(child.login, logs_data, targets)
        if sheet_url:
            return redirect(sheet_url)
        else:
            flash('Ошибка при создании отчета в Google Sheets', 'error')
            return redirect(url_for('parent_children'))

    # Подготавливаем сводные данные
    total_cal = sum(safe_float(l.calories) for l in logs)
    total_prot = sum(safe_float(l.protein) for l in logs)
    total_fat = sum(safe_float(l.fat) for l in logs)
    total_carbs = sum(safe_float(l.carbs) for l in logs)

    # Формируем имя файла
    start_label = start.date().isoformat()
    end_label = now.date().isoformat()
    base_name = f"{child.login}_nutrition_{start_label}_to_{end_label}"

    # Инициализируем html_parts по умолчанию
    html_parts = []

    if fmt == 'word' or fmt == 'doc':
        # Генерируем HTML-отчёт (вернём как .doc — Word откроет HTML).
        # Отчёт разбит по дням; для каждой даты перечисляем блюда и суммируем КБЖУ.
        # Ячейки итогов за день подсвечиваем: красным если > цели, синим если < цели, зелёным в остальных случаях.
        from collections import defaultdict

        daily = defaultdict(list)
        for l in logs:
            try:
                d = l.created_at.date().isoformat()
            except Exception:
                # fallback если created_at не валиден
                d = str(getattr(l, 'created_at', ''))
            daily[d].append(l)

        html_parts = []
        html_parts.append(f"<h2>Отчёт питания за год — {child.login}</h2>")
        html_parts.append(f"<p>Период: {start_label} — {end_label}</p>")

        # Стиль таблицы
        html_parts.append('<style>table{border-collapse:collapse;width:100%;}th,td{border:1px solid #ccc;padding:6px;text-align:left;}th{background:#f5f5f5;}</style>')

        # Проходим по датам в порядке возрастания
        for date_key in sorted(daily.keys()):
            day_logs = daily[date_key]
            html_parts.append(f"<h3>{date_key}</h3>")
            html_parts.append('<table>')
            html_parts.append('<tr><th>Время</th><th>Блюдо</th><th>Калории</th><th>Белки</th><th>Жиры</th><th>Углеводы</th></tr>')

            day_total = {'calories': 0.0, 'protein': 0.0, 'fat': 0.0, 'carbs': 0.0}
            for l in sorted(day_logs, key=lambda x: getattr(x, 'created_at', '')):
                time_str = ''
                try:
                    time_str = l.created_at.time().strftime('%H:%M')
                except Exception:
                    time_str = ''
                cal = float(l.calories or 0)
                prot = float(l.protein or 0)
                fat = float(l.fat or 0)
                carbs = float(l.carbs or 0)
                day_total['calories'] += cal
                day_total['protein'] += prot
                day_total['fat'] += fat
                day_total['carbs'] += carbs
                html_parts.append(f"<tr><td>{time_str}</td><td>{l.name}</td><td>{round(cal,1)}</td><td>{round(prot,1)}</td><td>{round(fat,1)}</td><td>{round(carbs,1)}</td></tr>")

            # Сравниваем итоги с целями и подготавливаем стили
            def cell_style(value, target):
                try:
                    v = float(value)
                    t = float(target)
                except Exception:
                    return ''
                if v > t:
                    return 'background-color:#ffecec;'
                if v < t:
                    return 'background-color:#ecf5ff;'
                return 'background-color:#ecffec;'

            html_parts.append('<tr>')
            html_parts.append('<td colspan="2"><strong>Итого за день:</strong></td>')
            html_parts.append(f"<td style=\"{cell_style(day_total['calories'], targets['calories'])}\"><strong>{round(day_total['calories'],1)}</strong></td>")
            html_parts.append(f"<td style=\"{cell_style(day_total['protein'], targets['protein'])}\"><strong>{round(day_total['protein'],1)}</strong></td>")
            html_parts.append(f"<td style=\"{cell_style(day_total['fat'], targets['fat'])}\"><strong>{round(day_total['fat'],1)}</strong></td>")
            html_parts.append(f"<td style=\"{cell_style(day_total['carbs'], targets['carbs'])}\"><strong>{round(day_total['carbs'],1)}</strong></td>")
            html_parts.append('</tr>')

            # Отдельная строка: отклонение (пишем + при превышении, - при нехватке, 0 при норме)
            def dev_cell_html(value, target):
                try:
                    v = float(value)
                    t = float(target)
                except Exception:
                    return '<td>0</td>'
                d = round(v - t, 1)
                if d > 0:
                    return f'<td><span style="color:#c00">+{d}</span></td>'
                if d < 0:
                    return f'<td><span style="color:#06c">-{abs(d)}</span></td>'
                return '<td>0</td>'

            html_parts.append('<tr>')
            html_parts.append('<td colspan="2"><strong>Отклонение</strong></td>')
            html_parts.append(dev_cell_html(day_total['calories'], targets['calories']))
            html_parts.append(dev_cell_html(day_total['protein'], targets['protein']))
            html_parts.append(dev_cell_html(day_total['fat'], targets['fat']))
            html_parts.append(dev_cell_html(day_total['carbs'], targets['carbs']))
            html_parts.append('</tr>')

            # Показать нормы под таблицей
            html_parts.append('<tr>')
            html_parts.append('<td colspan="2">Норма:</td>')
            html_parts.append(f"<td>{round(targets['calories'],1)}</td>")
            html_parts.append(f"<td>{round(targets['protein'],1)}</td>")
            html_parts.append(f"<td>{round(targets['fat'],1)}</td>")
            html_parts.append(f"<td>{round(targets['carbs'],1)}</td>")
            html_parts.append('</tr>')

            html_parts.append('</table><br/>')

        html = '<html><head><meta charset="utf-8"></head><body>' + ''.join(html_parts) + '</body></html>'
        data = html.encode('utf-8')
        # Use send_file with a BytesIO to improve reverse-proxy / mobile compatibility
        bio_doc = io.BytesIO(data)
        bio_doc.seek(0)
        # Diagnostic log for mobile download issues
        try:
            app.logger.info(
                f"EXPORT DOC: user_id={session.get('user_id')} role={session.get('role')} remote={request.remote_addr} "
                f"UA={request.headers.get('User-Agent')[:200]} content_type=application/msword content_length={len(data)}")
        except Exception:
            app.logger.exception('Failed to log export doc info')
        # Write brief debug file with first bytes (hex) so we can inspect without system logs
        try:
            Path('logs').mkdir(parents=True, exist_ok=True)
            preview = binascii.hexlify(data[:200]).decode('ascii')
            with open(Path('logs') / 'export_debug.txt', 'a', encoding='utf-8') as df:
                df.write(f"{datetime.utcnow().isoformat()} EXPORT DOC student_id={student_id} user_id={session.get('user_id')} role={session.get('role')} remote={request.remote_addr} fmt=doc size={len(data)}\n")
                df.write(preview + "\n\n")
        except Exception:
            app.logger.exception('Failed to write export debug file for DOC')
        return send_file(bio_doc, mimetype='application/msword', as_attachment=True, download_name=f"{base_name}.doc")

    # Excel (.xlsx) export with styling
    if fmt in ('excel', 'xlsx'):
        try:
            # Lazy-import openpyxl so it's optional for users who don't need Excel export
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment
        except Exception:
            flash('Для экспорта в Excel требуется библиотека openpyxl. Установите её: pip install openpyxl', 'error')
            return redirect(url_for('parent_children'))

        from collections import defaultdict

        daily = defaultdict(list)
        for l in logs:
            try:
                d = l.created_at.date().isoformat()
            except Exception:
                d = str(getattr(l, 'created_at', ''))
            daily[d].append(l)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Nutrition'

        row = 1
        ws.cell(row=row, column=1, value=f"Отчёт питания за год — {child.login}")
        row += 2
        ws.cell(row=row, column=1, value=f"Период: {start_label} — {end_label}")
        row += 2

        red_fill = PatternFill(start_color='FFFFECEC', fill_type='solid')
        blue_fill = PatternFill(start_color='FFECF5FF', fill_type='solid')
        green_fill = PatternFill(start_color='FFECFFEC', fill_type='solid')
        bold = Font(bold=True)

        # Columns header template
        for date_key in sorted(daily.keys()):
            day_logs = daily[date_key]

            ws.cell(row=row, column=1, value=date_key)
            ws.cell(row=row, column=1).font = bold
            row += 1
            # table header
            headers = ['Время', 'Блюдо', 'Калории', 'Белки', 'Жиры', 'Углеводы']
            for col, h in enumerate(headers, start=1):
                c = ws.cell(row=row, column=col, value=h)
                c.font = bold
            row += 1

            day_total = {'calories': 0.0, 'protein': 0.0, 'fat': 0.0, 'carbs': 0.0}
            for l in sorted(day_logs, key=lambda x: getattr(x, 'created_at', '')):
                time_str = ''
                try:
                    time_str = l.created_at.time().strftime('%H:%M')
                except Exception:
                    time_str = ''
                cal = float(l.calories or 0)
                prot = float(l.protein or 0)
                fat = float(l.fat or 0)
                carbs = float(l.carbs or 0)

                ws.cell(row=row, column=1, value=time_str)
                ws.cell(row=row, column=2, value=l.name)
                ws.cell(row=row, column=3, value=round(cal,1))
                ws.cell(row=row, column=4, value=round(prot,1))
                ws.cell(row=row, column=5, value=round(fat,1))
                ws.cell(row=row, column=6, value=round(carbs,1))

                day_total['calories'] += cal
                day_total['protein'] += prot
                day_total['fat'] += fat
                day_total['carbs'] += carbs
                row += 1

            # Totals row
            ws.cell(row=row, column=1, value='Итого за день:')
            ws.cell(row=row, column=1).font = bold

            # write numeric totals into cells
            ws.cell(row=row, column=3, value=round(day_total['calories'], 1))
            ws.cell(row=row, column=4, value=round(day_total['protein'], 1))
            ws.cell(row=row, column=5, value=round(day_total['fat'], 1))
            ws.cell(row=row, column=6, value=round(day_total['carbs'], 1))

            # Prepare differences
            diff_cal = round(day_total['calories'] - targets['calories'], 1)
            diff_prot = round(day_total['protein'] - targets['protein'], 1)
            diff_fat = round(day_total['fat'] - targets['fat'], 1)
            diff_carbs = round(day_total['carbs'] - targets['carbs'], 1)

            # Отклонение — отдельная строка: +N (красный) при превышении, -N (синий) при нехватке, 0 при норме
            row += 1
            ws.cell(row=row, column=1, value='Отклонение')
            ws.cell(row=row, column=1).font = bold

            from openpyxl.styles import Font
            red_font = Font(color='00FF0000', bold=True)
            blue_font = Font(color='000000FF', bold=True)

            def write_dev(col, diff):
                if diff > 0:
                    c = ws.cell(row=row, column=col, value=f'+{diff}')
                    c.font = red_font
                elif diff < 0:
                    c = ws.cell(row=row, column=col, value=f'-{abs(diff)}')
                    c.font = blue_font
                else:
                    ws.cell(row=row, column=col, value='0')

            write_dev(3, diff_cal)
            write_dev(4, diff_prot)
            write_dev(5, diff_fat)
            write_dev(6, diff_carbs)

            row += 1
            # Norms row
            ws.cell(row=row, column=1, value='Норма:')
            ws.cell(row=row, column=3, value=round(targets['calories'],1))
            ws.cell(row=row, column=4, value=round(targets['protein'],1))
            ws.cell(row=row, column=5, value=round(targets['fat'],1))
            ws.cell(row=row, column=6, value=round(targets['carbs'],1))
            row += 2

        # Auto-adjust column widths (simple heuristic)
        for col in range(1, 7):
            max_len = 0
            for r in range(1, row):
                v = ws.cell(row=r, column=col).value
                if v is None:
                    continue
                l = len(str(v))
                if l > max_len:
                    max_len = l
            ws.column_dimensions[chr(64+col)].width = min(50, max(10, max_len + 2))

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    data = bio.getvalue()
    # Diagnostic log for mobile download issues
    try:
        app.logger.info(
            f"EXPORT XLSX: user_id={session.get('user_id')} role={session.get('role')} remote={request.remote_addr} "
            f"UA={request.headers.get('User-Agent')[:200]} content_type=application/vnd.openxmlformats-officedocument.spreadsheetml.sheet content_length={len(data)}")
    except Exception:
        app.logger.exception('Failed to log export xlsx info')
    # Write brief debug file with first bytes (hex) so we can inspect without system logs
    try:
        Path('logs').mkdir(parents=True, exist_ok=True)
        preview = binascii.hexlify(data[:200]).decode('ascii')
        with open(Path('logs') / 'export_debug.txt', 'a', encoding='utf-8') as df:
            df.write(f"{datetime.utcnow().isoformat()} EXPORT XLSX student_id={student_id} user_id={session.get('user_id')} role={session.get('role')} remote={request.remote_addr} fmt=xlsx size={len(data)}\n")
            df.write(preview + "\n\n")
    except Exception:
        app.logger.exception('Failed to write export debug file for XLSX')
    fileobj = io.BytesIO(data)
    fileobj.seek(0)
    return send_file(fileobj, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f"{base_name}.xlsx")

    # По умолчанию — CSV для Excel
    si = io.StringIO()
    writer = csv.writer(si)
    # BOM для Excel (UTF-8) — поможет корректно показать кириллицу в Excel на Windows
    bom = '\ufeff'
    writer.writerow(['Дата', 'Время', 'Блюдо', 'Ккал', 'Белки', 'Жиры', 'Углеводы'])
    for l in logs:
        writer.writerow([l.created_at.date().isoformat(), l.created_at.time().strftime('%H:%M:%S'), l.name, l.calories, l.protein, l.fat, l.carbs])
    writer.writerow(['Итого', '', '', round(total_cal,1), round(total_prot,1), round(total_fat,1), round(total_carbs,1)])
    data = bom + si.getvalue()
    bdata = data.encode('utf-8')
    resp = make_response(bdata)
    resp.headers['Content-Type'] = 'text/csv; charset=utf-8'
    resp.headers['Content-Disposition'] = f'attachment; filename="{base_name}.csv"'
    resp.headers['Content-Length'] = str(len(bdata))
    # Diagnostic log for mobile download issues
    try:
        app.logger.info(
            f"EXPORT CSV: user_id={session.get('user_id')} role={session.get('role')} remote={request.remote_addr} "
            f"UA={request.headers.get('User-Agent')[:200]} content_type={resp.headers.get('Content-Type')} content_length={resp.headers.get('Content-Length')}")
    except Exception:
        app.logger.exception('Failed to log export csv info')
    # Write brief debug file with first bytes (hex) so we can inspect without system logs
    try:
        Path('logs').mkdir(parents=True, exist_ok=True)
        preview = binascii.hexlify(bdata[:200]).decode('ascii')
        with open(Path('logs') / 'export_debug.txt', 'a', encoding='utf-8') as df:
            df.write(f"{datetime.utcnow().isoformat()} EXPORT CSV student_id={student_id} user_id={session.get('user_id')} role={session.get('role')} remote={request.remote_addr} fmt=csv size={len(bdata)}\n")
            df.write(preview + "\n\n")
    except Exception:
        app.logger.exception('Failed to write export debug file for CSV')
    return resp


@app.route('/parent/child/<int:student_id>/report')
@login_required(role='parent')
def child_report(student_id: int):
    """Показывает веб-страницу с отчётом по питанию для ребёнка (доступен родителю)."""
    try:
        parent_id = get_session_user_id()
    except ValueError:
        flash('Ошибка сессии. Пожалуйста, войдите снова.', 'error')
        return redirect(url_for('login'))

    child = Student.query.get_or_404(student_id)
    if child.parent_id != parent_id:
        flash('Доступ запрещён', 'error')
        return redirect(url_for('parent_children'))

    # Собираем логи за год
    now = datetime.utcnow()
    start = now - timedelta(days=365)
    logs = EatLog.query.filter(
        EatLog.student_id == child.id,
        EatLog.created_at >= start,
        EatLog.created_at <= now
    ).order_by(EatLog.created_at.asc()).all()

    # targets
    targets = {
        'calories': safe_float(child.calories or 2000),
        'protein': safe_float(child.protein or 75),
        'fat': safe_float(child.fat or 60),
        'carbs': safe_float(child.carbs or 250)
    }

    # Group by date
    from collections import defaultdict
    daily = defaultdict(list)
    for l in logs:
        try:
            d = l.created_at.date().isoformat()
        except Exception:
            d = str(getattr(l, 'created_at', ''))
        daily[d].append(l)

    # prepare summary per day
    day_summaries = {}
    for date_key, day_logs in daily.items():
        consumed = get_nutrition_summary(day_logs)
        remaining = calculate_remaining_nutrients(targets, consumed)
        day_summaries[date_key] = {'logs': day_logs, 'consumed': consumed, 'remaining': remaining}

    return render_template('nutrition_report.html', child=child, day_summaries=day_summaries, targets=targets, start=start.date(), end=now.date())


@app.route('/parent/child/<int:student_id>/photo_analyze', methods=['GET', 'POST'])
@login_required(role='parent')
def photo_analyze_child(student_id: int):
    """Позволяет родителю загрузить фото и проанализировать его для конкретного ребёнка."""
    try:
        parent_id = get_session_user_id()
    except ValueError:
        flash('Ошибка сессии. Пожалуйста, войдите снова.', 'error')
        return redirect(url_for('login'))

    child = Student.query.get_or_404(student_id)
    if child.parent_id != parent_id:
        flash('Доступ запрещён', 'error')
        return redirect(url_for('parent_children'))

    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Файл не был отправлен. Пожалуйста, выберите файл.')
            return redirect(request.url)
        file = request.files['file']
        if not file or not file.filename:
            flash('Файл не был выбран.')
            return redirect(request.url)
        if file and allowed_file(file.filename):
            fname = secure_filename(file.filename or '')
            import time
            fname = f"{int(time.time())}_{fname}"
            filepath = str(UPLOAD_FOLDER / fname)
            try:
                file.save(filepath)
            except Exception as e:
                flash(f"Ошибка при сохранении файла: {e}")
                return redirect(request.url)
            app.logger.info(f'Running image analysis for file: {filepath}')
            nutrition_data = None
            data_url = None
            try:
                try:
                    if _ensure_analyze_image_loaded() and app._analyze_image:
                        nutrition_data = app._analyze_image(filepath)
                    else:
                        app.logger.warning('Image analysis disabled or failed to load; skipping analysis')
                        nutrition_data = None
                except Exception as e:
                    app.logger.exception(f'Unexpected exception from analyze_image_with_gemini: {e}')
                    nutrition_data = None

                # Read file bytes and build data URL so we can safely remove the file
                try:
                    import base64
                    with open(filepath, 'rb') as f:
                        b = f.read()
                    suffix = Path(filepath).suffix.lower().lstrip('.')
                    mime = 'jpeg' if suffix in ('jpg', 'jpeg') else suffix or 'octet-stream'
                    b64 = base64.b64encode(b).decode('ascii')
                    data_url = f'data:image/{mime};base64,{b64}'
                except Exception:
                    data_url = None

            finally:
                # Try to delete the uploaded file in all cases to avoid accumulation
                try:
                    p = Path(filepath)
                    if p.exists():
                        p.unlink()
                        app.logger.info(f'Removed uploaded file: {filepath}')
                except Exception:
                    app.logger.exception(f'Failed to remove uploaded file: {filepath}')

            if nutrition_data:
                # Результат получен (даже если это "No food detected")
                app.logger.info(f'Image analysis succeeded for file: {filepath} -> {nutrition_data.get("name") if isinstance(nutrition_data, dict) else "<non-dict>"}')
                return render_template('photo_analyze.html', filename=None, data=nutrition_data, data_url=data_url, target_student_id=student_id, is_authorized=True)
            else:
                # Результат = None, это реальная ошибка (ключ, импорт, сетевая ошибка)
                app.logger.warning(f'Image analysis returned None for file: {filepath}')
                flash('Не удалось проанализировать изображение. Возможно, файл поврежден или API временно недоступен.')
                return render_template('photo_analyze.html', filename=None, data=None, data_url=data_url, target_student_id=student_id, is_authorized=True)
        else:
            flash('Недопустимый тип файла. Разрешены: png, jpg, jpeg, gif, webp.')
            return redirect(request.url)

    return render_template('photo_analyze.html', filename=None, data=None, target_student_id=student_id, is_authorized=True)

from forms import FoodForm

@app.route('/add_food', methods=['GET', 'POST'])
@login_required()
def add_food():
    # Только ученики могут добавлять еду
    if session.get('role') != 'student':
        flash('Доступ запрещён', 'error')
        return redirect(url_for('login'))
        
    form = FoodForm()
    
    if request.method == 'POST':
        try:
            # Получаем данные из формы
            name = request.form.get('name')
            calories = safe_float(request.form.get('calories'))
            protein = safe_float(request.form.get('protein'))
            fat = safe_float(request.form.get('fat'))
            carbs = safe_float(request.form.get('carbs'))
            
            if not name or calories < 0 or protein < 0 or fat < 0 or carbs < 0:
                flash('Пожалуйста, заполните все поля корректными значениями', 'error')
                return render_template('add_food.html', form=form)
            
            # Обработка изображения
            file = request.files.get('image_file')
            image = request.form.get('image')  # URL изображения
            
            image_filename = None
            if file and file.filename:
                image_filename = save_image(file)
            elif image:
                image_filename = image
            
            # Создание новой записи о еде
            new_food = Eat(
                name=name,
                calories=calories,
                protein=protein,
                fat=fat,
                carbs=carbs,
                type='normal',
                image=image_filename
            )
            
            # Добавление штрихкода если есть
            barcode = request.form.get('barcode')
            if barcode:
                new_food.barcode = barcode.strip()
                
            # Сохранение в базу
            db.session.add(new_food)
            if safe_commit():
                app.logger.info(f'Food added: {new_food.name}')
                flash(f"Продукт {new_food.name} добавлен!", 'success')
                return redirect(url_for('eat'))
                
        except Exception as e:
            app.logger.error(f'Error adding food: {str(e)}')
            flash('Ошибка при добавлении продукта', 'error')
            db.session.rollback()

    return render_template('add_food.html', form=form)

@app.route('/calorie_calculator')
def calorie_calculator():
    return render_template('calorie_calculator.html')


# ---------------- Анализ фото еды ----------------
@app.route('/photo_analyze', methods=['GET', 'POST'])
def photo_analyze():
    """Анализ фото еды доступен всем пользователям"""
    # Проверяем, авторизован ли пользователь для показа кнопки добавления в дневник
    is_authorized = 'user_id' in session and session.get('role') in ('student', 'parent')
    
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Файл не был отправлен. Пожалуйста, выберите файл.')
            return redirect(request.url)
            
        file = request.files['file']
        if not file or not file.filename:
            flash('Файл не был выбран.')
            return redirect(request.url)
            
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename or '')
            filepath = str(UPLOAD_FOLDER / filename)
            
            try:
                file.save(filepath)
            except Exception as e:
                flash(f"Ошибка при сохранении файла: {e}")
                return redirect(request.url)
            # Используем функцию анализа изображения из анализатора-питания-по-фото
            app.logger.info(f'Running image analysis for file: {filepath}')
            nutrition_data = None
            data_url = None
            try:
                try:
                    if _ensure_analyze_image_loaded() and app._analyze_image:
                        nutrition_data = app._analyze_image(filepath)
                    else:
                        app.logger.warning('Image analysis disabled or module failed to load; skipping analysis')
                        nutrition_data = None
                except Exception as e:
                    app.logger.exception(f'Unexpected exception from analyze_image_with_gemini: {e}')
                    nutrition_data = None

                # Read file bytes and build data URL so we can safely remove the file
                try:
                    import base64
                    with open(filepath, 'rb') as f:
                        b = f.read()
                    suffix = Path(filepath).suffix.lower().lstrip('.')
                    mime = 'jpeg' if suffix in ('jpg', 'jpeg') else suffix or 'octet-stream'
                    b64 = base64.b64encode(b).decode('ascii')
                    data_url = f'data:image/{mime};base64,{b64}'
                except Exception:
                    data_url = None

            finally:
                # Try to delete the uploaded file in all cases to avoid accumulation
                try:
                    p = Path(filepath)
                    if p.exists():
                        p.unlink()
                        app.logger.info(f'Removed uploaded file: {filepath}')
                except Exception:
                    app.logger.exception(f'Failed to remove uploaded file: {filepath}')

            if nutrition_data:
                # Результат получен (даже если это "No food detected")
                app.logger.info(f'Image analysis succeeded for file: {filepath} -> {nutrition_data.get("name") if isinstance(nutrition_data, dict) else "<non-dict>"}')
                return render_template('photo_analyze.html', filename=None, data=nutrition_data, data_url=data_url, is_authorized=is_authorized)
            else:
                # Результат = None, это реальная ошибка (ключ, импорт, сетевая ошибка)
                app.logger.warning(f'Image analysis returned None for file: {filepath}')
                flash('Не удалось проанализировать изображение. Возможно, файл поврежден или API временно недоступен.')
                return render_template('photo_analyze.html', filename=None, data=None, data_url=data_url, is_authorized=is_authorized)
        else:
            flash('Недопустимый тип файла. Разрешены: png, jpg, jpeg, gif, webp.')
            return redirect(request.url)

    return render_template('photo_analyze.html', filename=None, data=None, is_authorized=is_authorized)

@app.route('/add_analyzed_food', methods=['POST'])
@login_required(role=['student', 'parent'])
def add_analyzed_food():
    """Добавляет проанализированную еду в дневник питания (только для авторизованных пользователей)"""
    try:
        # Проверяем базовые права: должен быть либо ученик, либо родитель
        role = session.get('role')
        if role not in ('student', 'parent'):
            flash('Доступ запрещён', 'error')
            return redirect(url_for('login'))
        calories = float(request.form.get('calories', 0))
        protein = float(request.form.get('protein', 0))
        fat = float(request.form.get('fat', 0))
        carbs = float(request.form.get('carbs', 0))
        name = request.form.get('name', 'Неизвестное блюдо')
        serving_size = request.form.get('serving_size', '100г')

        # Опционально: student_id — если родитель добавляет еду для ребёнка
        student_id = request.form.get('student_id')
        target_student_id = None
        if student_id:
            try:
                target_student_id = int(student_id)
            except Exception:
                target_student_id = None

        # Если передан target_student_id — проверяем права родителя
        if target_student_id:
            # must be parent and owner of child
            if session.get('role') != 'parent':
                flash('Недостаточно прав для добавления еды другому пользователю', 'error')
                return redirect(url_for('parent_children'))
            parent_id = get_session_user_id()
            child = Student.query.get(target_student_id)
            if not child or child.parent_id != parent_id:
                flash('Ребёнок не найден или доступ запрещён', 'error')
                return redirect(url_for('parent_children'))

            # Логируем только в базу для ребёнка — не меняем сессию родителя
            try:
                log = EatLog(
                    student_id=target_student_id,
                    food_id=None,
                    name=name,
                    calories=round(calories, 1),
                    protein=round(protein, 1),
                    fat=round(fat, 1),
                    carbs=round(carbs, 1)
                )
                db.session.add(log)
                if safe_commit():
                    flash(f"Блюдо {name} добавлено в дневник ребёнка {child.login}!")
                else:
                    flash('Ошибка при сохранении в дневник', 'error')
            except Exception:
                db.session.rollback()

            return redirect(url_for('parent_children'))

        # Иначе — стандартный путь для ученика: обновляем сессию текущего пользователя
        session["calories"] = round(session.get("calories", 0) - calories, 1)
        session["protein"] = round(session.get("protein", 0) - protein, 1)
        session["fat"] = round(session.get("fat", 0) - fat, 1)
        session["carbs"] = round(session.get("carbs", 0) - carbs, 1)
        if "eaten" not in session:
            session["eaten"] = []
        session["eaten"].append(f"{name} ({serving_size})")
        session.modified = True

        # Логируем в базу данных для текущего ученика
        try:
            log = EatLog(
                student_id=session.get('user_id'),
                food_id=None,  # это проанализированная еда, у нее нет id в базе
                name=name,
                calories=round(calories, 1),
                protein=round(protein, 1),
                fat=round(fat, 1),
                carbs=round(carbs, 1)
            )
            db.session.add(log)
            if safe_commit():
                flash(f"Блюдо {name} добавлено в дневник питания!")
            else:
                flash('Ошибка при сохранении в дневник', 'error')
        except Exception:
            db.session.rollback()
            flash('Ошибка при сохранении в базу данных', 'error')

    except ValueError:
        flash('Ошибка при обработке данных о питательности', 'error')
    except Exception as e:
        flash(f'Неожиданная ошибка: {str(e)}', 'error')

    return redirect(url_for('photo_analyze'))

# ---------------- API endpoints ----------------
# Удалено дублирующееся определение photo_analyze

@app.route('/api/v1/food/<int:food_id>')
@login_required()
def get_food(food_id):
    food = Eat.query.get_or_404(food_id)
    return jsonify({
        'id': food.id,
        'name': food.name,
        'calories': food.calories,
        'protein': food.protein,
        'fat': food.fat,
        'carbs': food.carbs,
        'type': food.type,
        'image': food.image
    })

# ---------------- Отчеты по питанию ----------------
@app.route('/example_nutrition_report')
@login_required()
def example_nutrition_report():
    """Показывает отчет по питанию для конкретного ученика"""
    if session.get('role') not in ('student', 'parent'):
        flash('Доступ запрещён', 'error')
        return redirect(url_for('login'))
        
    student_id = request.args.get('student_id')
    if student_id:
        # Проверка прав доступа к данным ученика
        if session.get('role') == 'parent':
            # Родитель может смотреть только своих детей
            child = Student.query.get_or_404(student_id)
            if child.parent_id != session.get('user_id'):
                flash('Доступ запрещён', 'error')
                return redirect(url_for('dashboard'))
    
    return render_template('example_nutrition_report.html')

# ---------------- API: cities / schools ----------------
    try:
        student = Student.query.get_or_404(student_id)
        
        # Собираем логи за год
        now = datetime.utcnow()
        start = now - timedelta(days=365)
        logs = EatLog.query.filter(
            EatLog.student_id == student.id,
            EatLog.created_at >= start,
            EatLog.created_at <= now
        ).order_by(EatLog.created_at.asc()).all()

        # Целевые показатели
        targets = {
            'calories': safe_float(student.calories or 2000),
            'protein': safe_float(student.protein or 75),
            'fat': safe_float(student.fat or 60),
            'carbs': safe_float(student.carbs or 250)
        }

        # Group by date
        daily = defaultdict(list)
        for l in logs:
            try:
                d = l.created_at.date().isoformat()
            except Exception:
                d = str(getattr(l, 'created_at', ''))
            daily[d].append(l)

        # prepare summary per day
        day_summaries = {}
        for date_key, day_logs in daily.items():
            consumed = get_nutrition_summary(day_logs)
            remaining = calculate_remaining_nutrients(targets, consumed)
            day_summaries[date_key] = {'logs': day_logs, 'consumed': consumed, 'remaining': remaining}

        return render_template('nutrition_report.html', 
                            child=student,  # используем тот же шаблон что и для родителя
                            day_summaries=day_summaries,
                            targets=targets,
                            start=start.date(),
                            end=now.date())
                            
    except Exception as e:
        app.logger.error(f'Error in teacher_student_report: {str(e)}')
        flash('Произошла ошибка при формировании отчета', 'error')
        return redirect(url_for('teacher_students'))



# ---------------- API: cities / schools ----------------
@app.route('/api/cities')
def api_cities():
    cities = City.query.order_by(City.name).all()
    return jsonify([{'id': c.id, 'name': c.name} for c in cities])


@app.route('/api/schools')
def api_schools():
    # query param: city (name)
    city = request.args.get('city')
    if not city:
        return jsonify([])
    c = City.query.filter_by(name=city).first()
    if not c:
        return jsonify([])
    schools = School.query.filter_by(city_id=c.id).order_by(School.name).all()
    return jsonify([{'id': s.id, 'name': s.name} for s in schools])

@app.route('/api/grades')
def api_grades():
    school = request.args.get('school')
    if not school:
        return jsonify([])
    s = School.query.filter_by(name=school).first()
    if not s:
        return jsonify([])
    grades = Grade.query.filter_by(school_id=s.id).order_by(Grade.name).all()
    return jsonify([{'id': g.id, 'name': g.name} for g in grades])


# ---------------- Admin management ----------------

@app.route('/admin_autologin')
def admin_autologin():
    """Endpoint для автоматического входа администратора."""
    try:
        # Проверяем наличие флешки
        has_flash = check_admin_flash_drive()
        app.logger.info(f"admin_autologin: checking flash drive: {has_flash}")
        
        if not has_flash:
            flash('Для входа администратора необходимо подключить флешку', 'error')
            return redirect(url_for('login'))

        # Если уже админ — перенаправляем на панель
        if session.get('role') == 'admin':
            return redirect(url_for('admin_dashboard'))

        # Пытаемся найти мастер-админа
        admin = Admin.query.filter_by(is_master=True).first()

        # Если нет мастер-админа, берем первого админа
        if not admin:
            admin = Admin.query.first()

        if admin:
            session['user_id'] = admin.id
            session['role'] = 'admin'
            session['is_master_admin'] = admin.is_master
            session.permanent = True
            app.logger.info(f"Автовход под админом: {admin.login}")
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Администратор не найден в системе', 'error')
            return redirect(url_for('index'))
    except Exception as e:
        app.logger.exception('Ошибка при автовходе администратора')
        flash('Произошла ошибка при входе в систему', 'error')
        return redirect(url_for('index'))

@app.route('/admin')
@login_required(role='admin')
def admin_dashboard():
    """Главная панель администратора"""
    cities = City.query.order_by(City.name).all()
    schools = School.query.join(City).order_by(City.name, School.name).all()
    grades = Grade.query.order_by(Grade.name).all()
    
    return render_template(
        'admin_panel.html',
        cities=cities,
        schools=schools,
        grades=grades
    )

@app.route('/admin/create', methods=['GET', 'POST'])
@login_required(role='admin')
def admin_create():
    """Создание нового администратора (только для master admin)"""
    if not session.get('is_master_admin'):
        flash('Только главный администратор может создавать новых администраторов', 'error')
        return redirect(url_for('admin_panel'))

    if request.method == 'POST':
        login = request.form.get('login')
        password = request.form.get('password')
        
        if not login or not password:
            flash('Необходимо указать логин и пароль', 'error')
            return redirect(url_for('admin_create'))
            
        try:
            admin = Admin(
                login=login,
                password=generate_password_hash(password),
                is_master=False,
                created_by=session.get('user_id')
            )
            db.session.add(admin)
            db.session.commit()
            flash('Администратор успешно создан', 'success')
            return redirect(url_for('admin_panel'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Ошибка при создании администратора: {str(e)}', 'error')
            
    return render_template('admin_create.html')

# ---------------- Admin: manage cities/schools/grades ----------------
@app.route('/admin/cities', methods=['GET', 'POST'])
@login_required(role='admin')
def admin_cities():
    if request.method == 'POST':
        name = request.form.get('name')
        if name:
            try:
                db.session.add(City(name=name.strip()))
                db.session.commit()
                flash('Город добавлен', 'success')
            except Exception:
                db.session.rollback()
                flash('Ошибка при добавлении города', 'error')
        # Остаёмся на вкладке "Города"
        return redirect(url_for('admin_dashboard', tab='cities'))
    cities = City.query.order_by(City.name).all()
    return render_template('admin_panel.html', cities=cities, schools=School.query.join(City).order_by(City.name, School.name).all(), grades=Grade.query.order_by(Grade.name).all())


@app.route('/admin/cities/delete/<int:city_id>', methods=['POST'])
@login_required(role='admin')
def admin_delete_city(city_id):
    c = City.query.get_or_404(city_id)
    try:
        # cascade manual: delete schools in this city first
        School.query.filter_by(city_id=c.id).delete()
        db.session.delete(c)
        db.session.commit()
        flash('Город удалён', 'success')
    except Exception:
        db.session.rollback()
        flash('Ошибка при удалении города', 'error')
    return redirect(url_for('admin_cities'))


@app.route('/admin/cities/edit/<int:city_id>', methods=['POST'])
@login_required(role='admin')
def admin_edit_city(city_id):
    c = City.query.get_or_404(city_id)
    new_name = request.form.get('name')
    if new_name:
        try:
            c.name = new_name.strip()
            db.session.commit()
            flash('Город обновлён', 'success')
        except Exception:
            db.session.rollback()
            flash('Ошибка при обновлении города', 'error')
    return redirect(url_for('admin_cities'))


@app.route('/admin/schools', methods=['GET', 'POST'])
@login_required(role='admin')
def admin_schools():
    if request.method == 'POST':
        name = request.form.get('name')
        city_id = request.form.get('city_id')
        if name and city_id:
            try:
                db.session.add(School(name=name.strip(), city_id=int(city_id)))
                db.session.commit()
                flash('Школа добавлена', 'success')
            except Exception:
                db.session.rollback()
                flash('Ошибка при добавлении школы', 'error')
        # Остаёмся на вкладке "Школы"
        return redirect(url_for('admin_dashboard', tab='schools'))
    cities = City.query.order_by(City.name).all()
    schools = School.query.join(City).order_by(City.name, School.name).all()
    return render_template('admin_panel.html', cities=cities, schools=schools, grades=Grade.query.order_by(Grade.name).all())


@app.route('/admin/schools/delete/<int:school_id>', methods=['POST'])
@login_required(role='admin')
def admin_delete_school(school_id):
    s = School.query.get_or_404(school_id)
    try:
        db.session.delete(s)
        db.session.commit()
        flash('Школа удалена', 'success')
    except Exception:
        db.session.rollback()
        flash('Ошибка при удалении школы', 'error')
    return redirect(url_for('admin_schools'))


@app.route('/admin/schools/edit/<int:school_id>', methods=['POST'])
@login_required(role='admin')
def admin_edit_school(school_id):
    s = School.query.get_or_404(school_id)
    new_name = request.form.get('name')
    new_city_id = request.form.get('city_id')
    if new_name and new_city_id:
        try:
            s.name = new_name.strip()
            s.city_id = int(new_city_id)
            db.session.commit()
            flash('Школа обновлена', 'success')
        except Exception:
            db.session.rollback()
            flash('Ошибка при обновлении школы', 'error')
    return redirect(url_for('admin_schools'))


@app.route('/admin/grades', methods=['GET', 'POST'])
@login_required(role='admin')
def admin_grades():
    if request.method == 'POST':
        name = request.form.get('name')
        school_id = request.form.get('school_id')
        if name and school_id:
            try:
                db.session.add(Grade(name=name.strip(), school_id=int(school_id)))
                db.session.commit()
                flash('Класс добавлен', 'success')
            except Exception:
                db.session.rollback()
                flash('Ошибка при добавлении класса', 'error')
        return redirect(url_for('admin_dashboard', tab='grades'))
    grades = Grade.query.order_by(Grade.name).all()
    schools = School.query.order_by(School.name).all()
    return render_template('admin_panel.html', cities=City.query.order_by(City.name).all(), schools=schools, grades=grades)


@app.route('/admin/grades/delete/<int:grade_id>', methods=['POST'])
@login_required(role='admin')
def admin_delete_grade(grade_id):
    g = Grade.query.get_or_404(grade_id)
    try:
        db.session.delete(g)
        db.session.commit()
        flash('Класс удалён', 'success')
    except Exception:
        db.session.rollback()
        flash('Ошибка при удалении класса', 'error')
    return redirect(url_for('admin_grades'))


@app.route('/admin/grades/edit/<int:grade_id>', methods=['POST'])
@login_required(role='admin')
def admin_edit_grade(grade_id):
    g = Grade.query.get_or_404(grade_id)
    new_name = request.form.get('name')
    if new_name:
        try:
            g.name = new_name.strip()
            db.session.commit()
            flash('Класс обновлён', 'success')
        except Exception:
            db.session.rollback()
            flash('Ошибка при обновлении класса', 'error')
    return redirect(url_for('admin_grades'))


# Роуты для анализа по фото удалены — функциональность определения по фото отключена.
# Если нужно полностью удалить связанные шаблоны и ссылки в UI, выполните дополнительные правки.

@app.route('/about')
def about():
    return render_template('about.html')


# ---- Debug helper: показать текущую сессию (полезно при локальной разработке) ----
@app.route('/_debug/session')
def _debug_session():
    # Возвращаем только полезные ключи
    return jsonify({
        'language': session.get('language'),
        'user_id': session.get('user_id'),
        'role': session.get('role')
    })

# API endpoints

@app.route('/api/v1/student/<int:student_id>/logs')
@login_required()
@cache.cached(timeout=60)
def get_student_logs(student_id):
    # Проверяем права доступа
    if session.get('role') == 'student' and session.get('user_id') != student_id:
        return jsonify({'error': 'Доступ запрещен'}), 403
    if session.get('role') == 'parent':
        parent = Parents.query.get(session.get('user_id'))
        if not parent or not Student.query.filter_by(id=student_id, parent_id=parent.id).first():
            return jsonify({'error': 'Доступ запрещен'}), 403
            
    start_of_day = datetime.combine(date.today(), datetime.min.time())
    logs = EatLog.query.filter(
        EatLog.student_id == student_id,
        EatLog.created_at >= start_of_day
    ).all()
    
    return jsonify([{
        'id': log.id,
        'name': log.name,
        'calories': log.calories,
        'protein': log.protein,
        'fat': log.fat,
        'carbs': log.carbs,
        'created_at': log.created_at.isoformat()
    } for log in logs])

# Error handlers
@app.errorhandler(404)
def not_found_error(error):
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Not found'}), 404
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Internal server error'}), 500
    return render_template('500.html'), 500





# ---------------- Запуск ----------------

# Инициализация базы данных и создание таблиц при запуске
with app.app_context():
    # Создаём таблицы, если их ещё нет
    db.create_all()

    # Проверяем и добавляем отсутствующие колонки для существующих таблиц
    try:
        # Проверяем barcode в таблице eat
        res = db.session.execute(text("PRAGMA table_info('eat')")).all()
        cols = [r[1] for r in res]
        if 'barcode' not in cols:
            try:
                db.session.execute(text("ALTER TABLE eat ADD COLUMN barcode VARCHAR(64)"))
                db.session.commit()
                print('Added missing column eat.barcode')
            except Exception as e:
                print('Failed to add barcode column:', e)

        # Проверяем is_active в таблице pack_items
        res = db.session.execute(text("PRAGMA table_info('pack_items')")).all()
        cols = [r[1] for r in res]
        if 'is_active' not in cols:
            try:
                db.session.execute(text("ALTER TABLE pack_items ADD COLUMN is_active BOOLEAN NOT NULL DEFAULT 1"))
                db.session.commit()
                print('Added missing column pack_items.is_active')
            except Exception as e:
                print('Failed to add is_active column:', e)

        # Проверяем новые колонки для Student (age, height, weight, activity, gender)
        res = db.session.execute(text("PRAGMA table_info('student')")).all()
        cols = [r[1] for r in res]
        print(f'[DB INIT] student table columns: {cols}')
        needed_cols = {
            'age': "REAL",
            'height': "REAL",
            'weight': "REAL",
            'activity': "REAL",
            'gender': "VARCHAR(10)"
        }
        for col_name, col_type in needed_cols.items():
            if col_name not in cols:
                try:
                    print(f'[DB INIT] Adding missing column student.{col_name}...')
                    db.session.execute(text(f"ALTER TABLE student ADD COLUMN {col_name} {col_type}"))
                    db.session.commit()
                    print(f'[DB INIT] Successfully added column student.{col_name}')
                except Exception as e:
                    print(f'[DB INIT] Failed to add {col_name} column: {e}')
            else:
                print(f'[DB INIT] Column student.{col_name} already exists')
    except Exception as e:
        # если что-то пошло не так с PRAGMA — не мешаем запуску
        print('Error checking/adding columns:', e)

# Запускаем приложение только при локальной разработке
if __name__ == "__main__":
    # Run locally when executed directly. In hosted environments (uwsgi/gunicorn)
    # the application should be served by the WSGI server instead.
    import os
    import socket

    def find_free_port(host: str, start_port: int = 5000, max_tries: int = 50) -> int:
        """Return first free port starting from start_port (tries max_tries)."""
        for p in range(start_port, start_port + max_tries):
            try:
                with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                    s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
                    s.bind((host, p))
                return p
            except OSError:
                continue
        raise RuntimeError(f'No free port found in range {start_port}..{start_port+max_tries-1}')

    host = os.environ.get('HOST', '0.0.0.0')
    try:
        requested_port = int(os.environ.get('PORT', '5000'))
    except Exception:
        requested_port = 5000

    try:
        port_to_use = find_free_port(host, requested_port, max_tries=100)
        if port_to_use != requested_port:
            app.logger.warning(f'Port {requested_port} in use, starting on {port_to_use} instead')
        app.run(host=host, port=port_to_use, debug=True)
    except Exception as e:
        app.logger.exception(f'Failed to start Flask dev server: {e}')
        raise
 